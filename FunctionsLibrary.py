# Import standard libraries
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import requests
import json
import statistics
import textwrap
from math import ceil
from datetime import datetime
from openpyxl import load_workbook, Workbook
from typing import List, Dict, Tuple, Optional, Callable, Any

# Import configuration
from config import (
    rules, data_req, ml_config, sim_config, llm_config, paths, events,
    display, confidence, get_event_encoding, is_valid_event, get_confidence_level
)

# ML Libraries for dual prediction system
try:
    import xgboost as xgb
    from sklearn.model_selection import train_test_split, cross_val_score
    from sklearn.metrics import mean_absolute_error, r2_score
    import joblib
    import os
    ML_AVAILABLE = True
except ImportError:
    ML_AVAILABLE = False
    print("Warning: XGBoost/scikit-learn not available. ML predictions disabled.")


# =============================================================================
# IMPORTS FROM REFACTORED MODULES (re-export for backward compatibility)
# =============================================================================

# Prediction functions
from woodchopping.predictions.llm import call_ollama
from woodchopping.predictions.baseline import (
    get_competitor_historical_times_flexible,
    get_event_baseline_flexible
)
from woodchopping.predictions.ml_model import (
    train_ml_model,
    predict_time_ml,
    perform_cross_validation,
    display_feature_importance
)
from woodchopping.predictions.ai_predictor import predict_competitor_time_with_ai
from woodchopping.predictions.prediction_aggregator import (
    get_all_predictions,
    select_best_prediction,
    generate_prediction_analysis_llm,
    display_dual_predictions
)

# Handicap calculation
from woodchopping.handicaps.calculator import calculate_ai_enhanced_handicaps

# Simulation functions
from woodchopping.simulation.monte_carlo import (
    simulate_single_race,
    run_monte_carlo_simulation
)
from woodchopping.simulation.visualization import (
    generate_simulation_summary,
    visualize_simulation_results
)
from woodchopping.simulation.fairness import (
    get_ai_assessment_of_handicaps,
    simulate_and_assess_handicaps
)

# UI functions
from woodchopping.ui.wood_ui import (
    wood_menu,
    select_wood_species,
    enter_wood_size_mm,
    enter_wood_quality,
    format_wood,
    select_event_code
)
from woodchopping.ui.competitor_ui import (
    select_all_event_competitors,
    competitor_menu,
    select_competitors_for_heat,
    view_heat_assignment,
    remove_from_heat
)
from woodchopping.ui.personnel_ui import personnel_management_menu
from woodchopping.ui.tournament_ui import (
    calculate_tournament_scenarios,
    distribute_competitors_into_heats,
    select_heat_advancers,
    generate_next_round,
    view_tournament_status,
    save_tournament_state,
    load_tournament_state,
    auto_save_state
)
from woodchopping.ui.handicap_ui import (
    view_handicaps_menu,
    view_handicaps
)

# File/sheet names (using config)
COMPETITOR_FILE = paths.EXCEL_FILE
COMPETITOR_SHEET = paths.COMPETITOR_SHEET
WOOD_FILE = paths.EXCEL_FILE
WOOD_SHEET = paths.WOOD_SHEET
RESULTS_FILE = paths.EXCEL_FILE
RESULTS_SHEET = paths.RESULTS_SHEET


# =============================================================================
# NON-MIGRATED FUNCTIONS (Excel I/O, Tournament Management, Data Validation)
# =============================================================================


# get_competitor_id_name_mapping
def get_competitor_id_name_mapping():
    
    try:
        df = pd.read_excel(COMPETITOR_FILE, sheet_name=COMPETITOR_SHEET)
        
        if df.empty:
            return {}, {}
        
        id_to_name = {}
        name_to_id = {}
        
        for _, row in df.iterrows():
            comp_id = str(row.get('CompetitorID', '')).strip()
            name = str(row.get('Name', '')).strip()
            
            if comp_id and name:
                id_to_name[comp_id] = name
                name_to_id[name.lower()] = comp_id
        
        return id_to_name, name_to_id
        
    except Exception as e:
        print(f"Error loading competitor ID/name mapping: {e}")
        return {}, {}


# load_competitors_df
def load_competitors_df() -> pd.DataFrame:
    """
    Load the competitor roster from Excel into a DataFrame.

    Returns:
        DataFrame with competitor information (name, country, ID, state, gender)
    """
    try:
        df = pd.read_excel(COMPETITOR_FILE, sheet_name=COMPETITOR_SHEET)

        # Standardize column names
        column_mapping = {
            'Name': 'competitor_name',
            'Country': 'competitor_country',
            'CompetitorID': 'competitor_id',
            'State/Province': 'state_province',
            'Gender': 'gender'
        }
        df = df.rename(columns=column_mapping)

        if df.empty:
            print("No competitors found in Excel. Please add competitors first.")
            return pd.DataFrame(columns=["competitor_name", "competitor_country"])

        print(f"Roster loaded successfully from Excel. Found {len(df)} competitors.")
        return df
        
    except FileNotFoundError:
        print(f"Excel file '{COMPETITOR_FILE}' not found. Creating new file.")
        # Create the Excel file with proper sheets
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        ws = wb.create_sheet(COMPETITOR_SHEET)
        ws.append(["CompetitorID", "Name", "Country", "State/Province", "Gender"])
        wb.save(COMPETITOR_FILE)
        wb.close()
        return pd.DataFrame(columns=["competitor_name", "competitor_country"])
    except Exception as e:
        print(f"Error loading roster from Excel: {e}")
        print(f"Looking for sheet: {COMPETITOR_SHEET}")
        return pd.DataFrame(columns=["competitor_name", "competitor_country"])



# load_wood_data
def load_wood_data():
    try:
        df = pd.read_excel(WOOD_FILE, sheet_name=WOOD_SHEET)
        return df
    except Exception as e:
        print(f"Error loading wood data: {e}")
        return pd.DataFrame(columns=["species", "multiplier"])



# load_results_df
def load_results_df():
    """Load the Results sheet as a DataFrame; returns empty DataFrame if missing."""
    try:
        df = pd.read_excel(RESULTS_FILE, sheet_name=RESULTS_SHEET)
        df.columns = [str(c).strip().lower() for c in df.columns]

        # Map Excel column names to expected names
        column_mapping = {
            'competitorid': 'competitor_id',
            'time (seconds)': 'raw_time',
            'size (mm)': 'size_mm',
            'species code': 'species',
            'round': 'round_name',        # NEW - for tournament support
            'heatid': 'heat_id'            # NEW - for tournament support
        }
        df = df.rename(columns=column_mapping)

        # Convert competitor IDs to names
        if 'competitor_id' in df.columns:
            id_to_name, _ = get_competitor_id_name_mapping()
            df['competitor_name'] = df['competitor_id'].apply(
                lambda x: id_to_name.get(str(x).strip(), str(x))
            )

        return df
    except Exception:
        return pd.DataFrame(columns=["event", "competitor_name", "species", "size_mm",
                                    "quality", "raw_time", "heat_id", "round_name", "timestamp"])



# add_competitor_with_times
def add_competitor_with_times():
    """Add a new competitor to roster and prompt for historical times"""
    try:
        # Get competitor basic info
        print("\n--- Add New Competitor ---")
        name = input("Enter competitor name: ").strip()
        if not name:
            print("Name cannot be empty.")
            return load_competitors_df()
        
        country = input("Enter competitor country: ").strip()
        state = input("Enter state/province (optional): ").strip()
        gender = input("Enter gender (M/F, optional): ").strip().upper()
        
        # Add to competitors sheet
        try:
            wb = load_workbook(COMPETITOR_FILE)
        except FileNotFoundError:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
        
        if COMPETITOR_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(COMPETITOR_SHEET)
            ws.append(["CompetitorID", "Name", "Country", "State/Province", "Gender"])
        else:
            ws = wb[COMPETITOR_SHEET]
            if ws.max_row < 1:
                ws.append(["CompetitorID", "Name", "Country", "State/Province", "Gender"])
        
        # Check for duplicate
        existing_names = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) > 1 and row[1]:  # Name is in column 2 (index 1)
                existing_names.add(str(row[1]).strip().lower())
        
        if name.lower() in existing_names:
            print("Competitor already exists in roster.")
            wb.close()
            return load_competitors_df()
        
        # Generate new CompetitorID
        new_id = f"C{str(ws.max_row).zfill(3)}"
        
        # Add competitor to sheet
        ws.append([new_id, name, country, state, gender])
        
        wb.save(COMPETITOR_FILE)
        wb.close()
        print(f"\n✓ {name} added to roster successfully with ID {new_id}")
        
        # Now prompt for historical times
        print("\n--- Enter Historical Competition Times ---")
        print("Minimum 3 times required for handicap calculation.")
        print("You can enter more than 3 if desired.")
        
        add_historical_times_for_competitor(name)
        
        return load_competitors_df()
        
    except Exception as e:
        print(f"Error adding competitor: {e}")
        return load_competitors_df()
    



# add_historical_times_for_competitor
def add_historical_times_for_competitor(competitor_name):
    """Prompt for and save historical times to results sheet
    Judge will be prompted to select whether each time is for SB or UH
    Judge will be prompted to enter the time, wood species, size, and date (optional)
    The program will store this data in the results sheet
    """
    times_added = 0
    
    while True:
        if times_added < 3:
            print(f"\n--- Historical Time Entry {times_added + 1} (minimum 3 required) ---")
        else:
            cont = input(f"\n{times_added} times entered. Add another? (y/n): ").strip().lower()
            if cont != 'y':
                break
            print(f"\n--- Historical Time Entry {times_added + 1} ---")
        
        # Get event type
        while True:
            event = input("Event type (SB for Standing Block, UH for Underhand): ").strip().upper()
            if event in ["SB", "UH"]:
                break
            print("Please enter SB or UH.")
        
        # Get time
        while True:
            time_str = input("Time in seconds (e.g., 45.3): ").strip()
            try:
                time_val = float(time_str)
                break
            except ValueError:
                print("Please enter a valid number.")
        
        # Get wood species
        species = input("Wood species: ").strip()
        
        # Get size
        while True:
            size_str = input("Wood diameter in mm: ").strip()
            try:
                size_val = float(size_str)
                break
            except ValueError:
                print("Please enter a valid number.")
        
        # Get quality (optional, default to 5)
        quality_str = input("Wood quality (0-10, press Enter for 5): ").strip()
        if quality_str:
            try:
                quality = int(quality_str)
                quality = max(0, min(10, quality))
            except:
                quality = 5
        else:
            quality = 5
        
        # Optional date
        date_str = input("Date (optional, format YYYY-MM-DD or press Enter to skip): ").strip()
        if not date_str:
            timestamp = datetime.now().isoformat(timespec="seconds")
        else:
            try:
                date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                timestamp = date_obj.isoformat(timespec="seconds")
            except:
                timestamp = datetime.now().isoformat(timespec="seconds")
        
        # Save to results sheet
        save_time_to_results(event, competitor_name, species, size_val, quality, time_val, 
                           f"Historical-{event}", timestamp)
        times_added += 1
        print(f"✓ Time #{times_added} saved successfully")
    
    if times_added >= 3:
        print(f"\n✓ {times_added} historical times added for {competitor_name}.")
    else:
        print(f"\n⚠ Warning: Only {times_added} times added. Minimum 3 recommended for handicap calculation.")




# save_time_to_results
def save_time_to_results(event, name, species, size, quality, time, heat_id, timestamp):
    """Helper to save a single time entry to results sheet"""
    try:
        # Convert name to ID
        _, name_to_id = get_competitor_id_name_mapping()
        competitor_id = name_to_id.get(str(name).strip().lower(), name)
        
        try:
            wb = load_workbook(RESULTS_FILE)
        except:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
        
        if RESULTS_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(RESULTS_SHEET)
            ws.append(["CompetitorID", "Event", "Time (seconds)", "Size (mm)", "Species Code", "Date"])
        else:
            ws = wb[RESULTS_SHEET]
            if ws.max_row == 0:
                ws.append(["CompetitorID", "Event", "Time (seconds)", "Size (mm)", "Species Code", "Date"])
        
        # Save with Excel's expected column names
        ws.append([competitor_id, event, time, size, species, timestamp])
        wb.save(RESULTS_FILE)
        wb.close()
        
    except Exception as e:
        print(f"Error saving time to results: {e}")


# validate_results_data
def validate_results_data(results_df: pd.DataFrame) -> Tuple[Optional[pd.DataFrame], List[str]]:
    """
    Validate and clean historical results data.

    Args:
        results_df: Raw results DataFrame

    Returns:
        Tuple of (cleaned_df, warnings_list). Returns (None, warnings) if validation fails.
    """
    warnings: List[str] = []

    if results_df is None or results_df.empty:
        return None, ["No data to validate"]

    df = results_df.copy()
    initial_count = len(df)

    # Check for required columns
    required_cols = ['competitor_name', 'event', 'raw_time', 'size_mm', 'species']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        warnings.append(f"Missing columns: {missing_cols}")
        return None, warnings

    # Remove invalid times (use config constants)
    invalid_times = df[
        (df['raw_time'] <= data_req.MIN_VALID_TIME_SECONDS) |
        (df['raw_time'] > data_req.MAX_VALID_TIME_SECONDS)
    ]
    if not invalid_times.empty:
        warnings.append(
            f"Removed {len(invalid_times)} records with impossible times "
            f"(<{data_req.MIN_VALID_TIME_SECONDS}s or >{data_req.MAX_VALID_TIME_SECONDS}s)"
        )
        df = df[
            (df['raw_time'] > data_req.MIN_VALID_TIME_SECONDS) &
            (df['raw_time'] <= data_req.MAX_VALID_TIME_SECONDS)
        ]

    # Remove invalid sizes (use config constants)
    invalid_sizes = df[
        (df['size_mm'] < data_req.MIN_DIAMETER_MM) |
        (df['size_mm'] > data_req.MAX_DIAMETER_MM)
    ]
    if not invalid_sizes.empty:
        warnings.append(
            f"Removed {len(invalid_sizes)} records with invalid diameters "
            f"(<{data_req.MIN_DIAMETER_MM}mm or >{data_req.MAX_DIAMETER_MM}mm)"
        )
        df = df[
            (df['size_mm'] >= data_req.MIN_DIAMETER_MM) &
            (df['size_mm'] <= data_req.MAX_DIAMETER_MM)
        ]

    # Check for missing competitor names
    missing_names = df[df['competitor_name'].isna() | (df['competitor_name'] == '')]
    if not missing_names.empty:
        warnings.append(f"Removed {len(missing_names)} records with missing competitor names")
        df = df[df['competitor_name'].notna() & (df['competitor_name'] != '')]

    # Check for invalid event codes (use config constants)
    invalid_events = df[~df['event'].isin(events.VALID_EVENTS)]
    if not invalid_events.empty:
        warnings.append(
            f"Removed {len(invalid_events)} records with invalid event codes "
            f"(must be {' or '.join(events.VALID_EVENTS)})"
        )
        df = df[df['event'].isin(events.VALID_EVENTS)]

    # Detect statistical outliers using IQR method (per event type)
    outliers_removed = 0
    for event in events.VALID_EVENTS:
        event_data = df[df['event'] == event]['raw_time']
        if len(event_data) > 10:  # Only check if we have enough data
            Q1 = event_data.quantile(0.25)
            Q3 = event_data.quantile(0.75)
            IQR = Q3 - Q1
            # Use config constant for IQR multiplier
            lower_bound = Q1 - data_req.OUTLIER_IQR_MULTIPLIER * IQR
            upper_bound = Q3 + data_req.OUTLIER_IQR_MULTIPLIER * IQR

            outliers = df[(df['event'] == event) &
                         ((df['raw_time'] < lower_bound) | (df['raw_time'] > upper_bound))]
            if not outliers.empty:
                outliers_removed += len(outliers)
                df = df[~((df['event'] == event) &
                         ((df['raw_time'] < lower_bound) | (df['raw_time'] > upper_bound)))]

    if outliers_removed > 0:
        warnings.append(
            f"Removed {outliers_removed} statistical outliers "
            f"(>{data_req.OUTLIER_IQR_MULTIPLIER}x IQR from median)"
        )

    final_count = len(df)
    if final_count < initial_count:
        warnings.append(
            f"Data cleaned: {initial_count} -> {final_count} records "
            f"({initial_count - final_count} removed)"
        )

    if final_count < data_req.MIN_ML_TRAINING_RECORDS_TOTAL:
        warnings.append(
            f"Warning: Only {final_count} valid records remaining "
            f"(need {data_req.MIN_ML_TRAINING_RECORDS_TOTAL}+ for ML training)"
        )

    return df, warnings




# engineer_features_for_ml
def engineer_features_for_ml(results_df, wood_df=None):
    """
    Engineer features for ML model from historical results.

    Args:
        results_df: DataFrame with historical results
        wood_df: DataFrame with wood properties (optional, will load if not provided)

    Returns:
        DataFrame with engineered features ready for training/prediction
    """
    if results_df is None or results_df.empty:
        return None

    # Load wood data if not provided
    if wood_df is None:
        wood_df = load_wood_data()

    # Create a copy to avoid modifying original
    df = results_df.copy()

    # Ensure required columns exist
    required_cols = ['competitor_name', 'event', 'raw_time', 'size_mm', 'species']
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        print(f"Warning: Missing required columns for ML: {missing}")
        return None

    # Remove invalid records
    df = df[df['raw_time'] > 0].copy()
    df = df[df['size_mm'] > 0].copy()

    if df.empty:
        return None

    # Feature 1: Event type encoding (SB=0, UH=1)
    df['event_encoded'] = df['event'].apply(lambda x: 0 if str(x).upper() == 'SB' else 1)

    # Feature 2: Competitor average time by event
    competitor_avg = df.groupby(['competitor_name', 'event'])['raw_time'].transform('mean')
    df['competitor_avg_time_by_event'] = competitor_avg

    # Feature 3: Competitor experience (count of past events)
    df['competitor_experience'] = df.groupby('competitor_name').cumcount() + 1

    # Feature 4: Size (already present as size_mm)

    # Feature 5 & 6: Join wood properties (janka_hardness, spec_gravity)
    if not wood_df.empty and 'speciesID' in wood_df.columns:
        # Create species code mapping - use speciesID to match with Results sheet species codes
        wood_properties = wood_df[['speciesID', 'janka_hard', 'spec_gravity']].copy()
        wood_properties = wood_properties.rename(columns={
            'speciesID': 'species',
            'janka_hard': 'wood_janka_hardness',
            'spec_gravity': 'wood_spec_gravity'
        })

        # Join wood properties with results
        df = df.merge(wood_properties, on='species', how='left')

        # Fill missing wood properties with median values (not inplace to avoid warning)
        median_janka = df['wood_janka_hardness'].median()
        median_spec_grav = df['wood_spec_gravity'].median()

        # Use fillna without inplace
        df['wood_janka_hardness'] = df['wood_janka_hardness'].fillna(median_janka if pd.notna(median_janka) else 2000)
        df['wood_spec_gravity'] = df['wood_spec_gravity'].fillna(median_spec_grav if pd.notna(median_spec_grav) else 0.37)
    else:
        # If wood data not available, use default values
        df['wood_janka_hardness'] = 2000  # Default medium hardness
        df['wood_spec_gravity'] = 0.37    # Default medium density

    return df




# validate_heat_data
def validate_heat_data(heat_assignment_df, wood_selection):
    if heat_assignment_df is None or heat_assignment_df.empty:
        print("\nNo competitors in heat assignment. Use Competitor Menu -> Select Competitors for Heat.")
        return False
    
    if not wood_selection.get("species") or not wood_selection.get("size_mm"):
        print("\nWood selection incomplete. Use Wood Menu to set species and size.")
        return False
    
    if not wood_selection.get("event"):
        print("\nEvent not selected. Use Wood Menu -> Select event (SB/UH) or Main Menu option 3.")
        return False
    
    return True

    # Calculate and display AI-enhanced handicap marks for the heat



# detect_results_sheet
def detect_results_sheet(wb):
    if RESULTS_SHEET in wb.sheetnames:
        return wb[RESULTS_SHEET]
    
    ws = wb.create_sheet(RESULTS_SHEET)
    ws.append(["CompetitorID", "Event", "Time (seconds)", "Size (mm)", "Species Code", "Date"])
    return ws

#Add the results to the 'Results' sheet in the Excel file
'''Prompts the user for: Event (SB/UH), species, wood, quality, and notes (could be heat ID or notes about finals/qualifier or simply the date)'''



# append_results_to_excel
def append_results_to_excel(heat_assignment_df, wood_selection, round_object=None, tournament_state=None):
    """Append heat results to Excel Results sheet.

    Args:
        heat_assignment_df (DataFrame): LEGACY - competitors in heat (for backward compatibility)
        wood_selection (dict): Wood characteristics
        round_object (dict): NEW - Round object from tournament system (optional)
        tournament_state (dict): NEW - Tournament state for context (optional)
    """

    # Determine if using new tournament system or legacy single-heat system
    if round_object is not None:
        # NEW TOURNAMENT SYSTEM
        competitors_list = round_object['competitors']
        round_name = round_object['round_name']
    else:
        # LEGACY SINGLE-HEAT SYSTEM
        if heat_assignment_df is None or heat_assignment_df.empty:
            print("No competitors in heat assignment.")
            return
        competitors_list = heat_assignment_df['competitor_name'].tolist()
        round_name = None

    event_code = wood_selection.get("event")
    if event_code not in ("SB", "UH"):
        print("Event not selected. Use Wood Menu -> Select event (SB/UH) or Main Menu option 3.")
        return

    species = wood_selection.get("species")
    size_mm = wood_selection.get("size_mm")
    quality = wood_selection.get("quality")

    # Generate HeatID
    if round_object and tournament_state:
        # NEW: Use round name and tournament context
        event_name = tournament_state.get('event_name', 'Event')
        heat_id = f"{event_code}-{event_name}-{round_name}".replace(" ", "-")
    else:
        # LEGACY: Prompt for Heat ID
        heat_id = input("Enter a Heat ID (e.g., SB-01-Qual or any short label): ").strip()
        if not heat_id:
            heat_id = f"{event_code}-{datetime.now().strftime('%Y%m%d-%H%M%S')}"

    # Get ID/name mapping
    _, name_to_id = get_competitor_id_name_mapping()

    rows_to_write = []
    times_collected = {}

    # STEP 1: Collect raw cutting times (for historical data)
    print("\n" + "=" * 70)
    print("STEP 1: RECORD CUTTING TIMES")
    print("=" * 70)
    print("Enter the raw cutting time for each competitor (from their mark to block severed)")
    print("Press Enter to skip a competitor\n")

    for name in competitors_list:
        s = input(f"  Cutting time for {name}: ").strip()

        if s == "":
            continue

        try:
            t = float(s)
            times_collected[name] = t
        except ValueError:
            print("    Invalid time; skipping this entry.")
            continue

    if not times_collected:
        print("\nNo results to record.")
        return

    # STEP 2: Record finish order (WHO FINISHED FIRST in real-time)
    print("\n" + "=" * 70)
    print("STEP 2: RECORD FINISH ORDER")
    print("=" * 70)
    print("Enter the finish position for each competitor (1 = finished first, 2 = second, etc.)")
    print("This is based on when they physically severed their block (handicap delays included)\n")

    finish_order = {}
    for name in times_collected.keys():
        while True:
            pos_str = input(f"  Finish position for {name}: ").strip()
            try:
                position = int(pos_str)
                if position < 1:
                    print("    Position must be 1 or greater")
                    continue
                finish_order[name] = position
                break
            except ValueError:
                print("    Invalid position; please enter a number")

    # Prepare rows for Excel and update round_object
    timestamp = datetime.now().isoformat(timespec="seconds")

    for name, time_val in times_collected.items():
        competitor_id = name_to_id.get(str(name).strip().lower(), name)

        # Include Round and HeatID columns
        rows_to_write.append([competitor_id, event_code, time_val, size_mm, species, timestamp, round_name or "", heat_id])

        # Store in round_object if using tournament system
        if round_object is not None:
            round_object['actual_results'][name] = time_val
            if 'finish_order' not in round_object:
                round_object['finish_order'] = {}
            round_object['finish_order'][name] = finish_order.get(name, 999)

    if not rows_to_write:
        print("No results to write.")
        return

    try:
        # Load or create workbook
        try:
            wb = load_workbook(RESULTS_FILE)
        except Exception:
            wb = Workbook()
            if "Sheet" in wb.sheetnames and RESULTS_SHEET not in wb.sheetnames:
                wb.remove(wb["Sheet"])

        ws = detect_results_sheet(wb)

        # Ensure header exists with Excel column names (NEW: includes Round and HeatID)
        if ws.max_row == 0:
            ws.append(["CompetitorID", "Event", "Time (seconds)", "Size (mm)", "Species Code", "Date", "Round", "HeatID"])

        # Append rows
        for r in rows_to_write:
            ws.append(r)

        wb.save(RESULTS_FILE)
        wb.close()
        print("Results appended to Excel successfully.")

        # NEW: Update round status if using tournament system
        if round_object is not None:
            round_object['status'] = 'in_progress'  # Mark as in progress (not completed until advancers selected)

    except Exception as e:
        print(f"Error appending results: {e}")
