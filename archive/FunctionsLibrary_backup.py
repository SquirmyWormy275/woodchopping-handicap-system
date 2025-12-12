1# Import standard libraries
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import requests
import json
import statistics
import textwrap
from math import ceil
from datetime import datetime
from openpyxl import load_workbook, Workbook
from typing import List, Dict, Tuple, Optional, Callable, Any

# Import configuration
from config import (
    rules, data_req, ml_config, sim_config, llm_config, paths, events,
    display, confidence, get_event_encoding, is_valid_event, get_confidence_level
)

# ML Libraries for dual prediction system
try:
    import xgboost as xgb
    from sklearn.model_selection import train_test_split, cross_val_score
    from sklearn.metrics import mean_absolute_error, r2_score
    import joblib
    import os
    ML_AVAILABLE = True
except ImportError:
    ML_AVAILABLE = False
    print("Warning: XGBoost/scikit-learn not available. ML predictions disabled.")


# File/sheet names (using config)
COMPETITOR_FILE = paths.EXCEL_FILE
COMPETITOR_SHEET = paths.COMPETITOR_SHEET
WOOD_FILE = paths.EXCEL_FILE
WOOD_SHEET = paths.WOOD_SHEET
RESULTS_FILE = paths.EXCEL_FILE
RESULTS_SHEET = paths.RESULTS_SHEET  

# CORE HELPER FUNCTIONS

"""Load competitor data and return two dictionaries: id_to_name and name_to_id
    All Ids and names are stored on the 'Competitor' sheet in the excel file."""

def get_competitor_id_name_mapping():
    
    try:
        df = pd.read_excel(COMPETITOR_FILE, sheet_name=COMPETITOR_SHEET)
        
        if df.empty:
            return {}, {}
        
        id_to_name = {}
        name_to_id = {}
        
        for _, row in df.iterrows():
            comp_id = str(row.get('CompetitorID', '')).strip()
            name = str(row.get('Name', '')).strip()
            
            if comp_id and name:
                id_to_name[comp_id] = name
                name_to_id[name.lower()] = comp_id
        
        return id_to_name, name_to_id
        
    except Exception as e:
        print(f"Error loading competitor ID/name mapping: {e}")
        return {}, {}


##Load the roster from Excel into a DataFrame.
def load_competitors_df() -> pd.DataFrame:
    """
    Load the competitor roster from Excel into a DataFrame.

    Returns:
        DataFrame with competitor information (name, country, ID, state, gender)
    """
    try:
        df = pd.read_excel(COMPETITOR_FILE, sheet_name=COMPETITOR_SHEET)

        # Standardize column names
        column_mapping = {
            'Name': 'competitor_name',
            'Country': 'competitor_country',
            'CompetitorID': 'competitor_id',
            'State/Province': 'state_province',
            'Gender': 'gender'
        }
        df = df.rename(columns=column_mapping)

        if df.empty:
            print("No competitors found in Excel. Please add competitors first.")
            return pd.DataFrame(columns=["competitor_name", "competitor_country"])

        print(f"Roster loaded successfully from Excel. Found {len(df)} competitors.")
        return df
        
    except FileNotFoundError:
        print(f"Excel file '{COMPETITOR_FILE}' not found. Creating new file.")
        # Create the Excel file with proper sheets
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        ws = wb.create_sheet(COMPETITOR_SHEET)
        ws.append(["CompetitorID", "Name", "Country", "State/Province", "Gender"])
        wb.save(COMPETITOR_FILE)
        wb.close()
        return pd.DataFrame(columns=["competitor_name", "competitor_country"])
    except Exception as e:
        print(f"Error loading roster from Excel: {e}")
        print(f"Looking for sheet: {COMPETITOR_SHEET}")
        return pd.DataFrame(columns=["competitor_name", "competitor_country"])


## Load wood speciess data from excel into a DataFrame.
def load_wood_data():
    try:
        df = pd.read_excel(WOOD_FILE, sheet_name=WOOD_SHEET)
        return df
    except Exception as e:
        print(f"Error loading wood data: {e}")
        return pd.DataFrame(columns=["species", "multiplier"])


def load_results_df():
    """Load the Results sheet as a DataFrame; returns empty DataFrame if missing."""
    try:
        df = pd.read_excel(RESULTS_FILE, sheet_name=RESULTS_SHEET)
        df.columns = [str(c).strip().lower() for c in df.columns]

        # Map Excel column names to expected names
        column_mapping = {
            'competitorid': 'competitor_id',
            'time (seconds)': 'raw_time',
            'size (mm)': 'size_mm',
            'species code': 'species',
            'round': 'round_name',        # NEW - for tournament support
            'heatid': 'heat_id'            # NEW - for tournament support
        }
        df = df.rename(columns=column_mapping)

        # Convert competitor IDs to names
        if 'competitor_id' in df.columns:
            id_to_name, _ = get_competitor_id_name_mapping()
            df['competitor_name'] = df['competitor_id'].apply(
                lambda x: id_to_name.get(str(x).strip(), str(x))
            )

        return df
    except Exception:
        return pd.DataFrame(columns=["event", "competitor_name", "species", "size_mm",
                                    "quality", "raw_time", "heat_id", "round_name", "timestamp"])


# ============================================================================
# TOURNAMENT MANAGEMENT FUNCTIONS - Multi-round heat generation system
# ============================================================================

def calculate_tournament_scenarios(num_stands, tentative_competitors):
    """Calculate three tournament format scenarios based on stands and competitor count.

    Args:
        num_stands (int): Available chopping stands
        tentative_competitors (int): Judge's estimate of competitor count

    Returns:
        dict: Contains all scenario options with format:
            {
                'single_heat': {scenario details},
                'heats_to_finals': {scenario details},
                'heats_to_semis_to_finals': {scenario details}
            }
    """

    # SCENARIO 0: Single Heat (Training/Testing)
    # Strategy: One heat only, perfect for practice/testing/small casual events
    scenario_0 = {
        'max_competitors': num_stands,
        'num_heats': 1,
        'num_semis': 0,
        'num_finals': 0,
        'advancers_per_heat': 0,  # No one advances, it's just one heat
        'total_blocks': num_stands,
        'description': (
            f"Single heat with up to {num_stands} competitors\n"
            f"  → Perfect for training, testing, or casual events\n"
            f"  → Results can still be saved to build historical data\n"
            f"  → No advancement rounds"
        )
    }

    # Calculate minimum heats needed
    min_heats = ceil(tentative_competitors / num_stands)

    # SCENARIO 1: Heats → Finals
    # Strategy: Use minimum heats, take top N from each to fill finals
    num_heats_s1 = max(min_heats, 2)  # At least 2 heats for a tournament
    max_competitors_s1 = num_heats_s1 * num_stands
    advancers_per_heat_s1 = num_stands // num_heats_s1

    # Ensure we can fill a final
    if advancers_per_heat_s1 * num_heats_s1 < num_stands:
        advancers_per_heat_s1 += 1

    total_blocks_s1 = tentative_competitors + num_stands

    scenario_1 = {
        'max_competitors': max_competitors_s1,
        'num_heats': num_heats_s1,
        'num_semis': 0,
        'num_finals': 1,
        'advancers_per_heat': advancers_per_heat_s1,
        'total_blocks': total_blocks_s1,
        'description': (
            f"{num_heats_s1} heats of {num_stands} (max {max_competitors_s1} competitors)\n"
            f"  → Top {advancers_per_heat_s1} from each heat advance\n"
            f"  → {num_stands}-person Final"
        )
    }

    # SCENARIO 2: Heats → Semis → Finals
    # Strategy: More heats to create semi-final round
    num_heats_s2 = max(min_heats + 2, 4)  # Add 2 more heats for semi tier
    max_competitors_s2 = num_heats_s2 * num_stands

    # Calculate semi-finals
    num_semis = 2
    semi_total = num_semis * num_stands
    advancers_per_heat_s2 = ceil(semi_total / num_heats_s2)
    advancers_per_semi = num_stands // num_semis

    total_blocks_s2 = tentative_competitors + semi_total + num_stands

    scenario_2 = {
        'max_competitors': max_competitors_s2,
        'num_heats': num_heats_s2,
        'num_semis': num_semis,
        'num_finals': 1,
        'advancers_per_heat': advancers_per_heat_s2,
        'advancers_per_semi': advancers_per_semi,
        'semi_total': semi_total,
        'total_blocks': total_blocks_s2,
        'description': (
            f"{num_heats_s2} heats of {num_stands} (max {max_competitors_s2} competitors)\n"
            f"  → Top {advancers_per_heat_s2} from each heat ({semi_total} total)\n"
            f"  → {num_semis} semi-finals of {num_stands}\n"
            f"  → Top {advancers_per_semi} from each semi\n"
            f"  → {num_stands}-person Final"
        )
    }

    return {
        'single_heat': scenario_0,
        'heats_to_finals': scenario_1,
        'heats_to_semis_to_finals': scenario_2
    }


def distribute_competitors_into_heats(all_competitors_df, handicap_results, num_stands, num_heats):
    """Distribute competitors into balanced heats using snake draft pattern.

    This algorithm ensures fair distribution of skill levels across all heats by:
    1. Sorting competitors by handicap mark (highest = front markers first)
    2. Using snake draft: forward then backward distribution pattern
    3. Automatically handling partial heats with appropriate advancement rules

    Args:
        all_competitors_df (DataFrame): All competitors in tournament
        handicap_results (list): Results from calculate_ai_enhanced_handicaps()
                                 Format: [{'name': ..., 'mark': ..., 'predicted_time': ...}, ...]
        num_stands (int): Competitors per heat (heat capacity)
        num_heats (int): Number of heats to create

    Returns:
        list: List of round_object dictionaries, one for each heat
    """

    # Sort competitors by mark (descending: highest mark first = front markers)
    sorted_competitors = sorted(handicap_results, key=lambda x: x['mark'], reverse=True)

    # Initialize empty heats
    heats = []
    for i in range(num_heats):
        heats.append({
            'round_type': 'heat',
            'round_number': i + 1,
            'round_name': f'Heat {i + 1}',
            'competitors': [],
            'competitors_df': pd.DataFrame(),
            'handicap_results': [],
            'actual_results': {},
            'advancers': [],
            'num_to_advance': 2,  # Default, will be adjusted for partial heats
            'status': 'pending'
        })

    # Snake draft distribution
    heat_index = 0
    direction = 1  # 1 = forward, -1 = backward

    for comp in sorted_competitors:
        heats[heat_index]['competitors'].append(comp['name'])
        heats[heat_index]['handicap_results'].append(comp)

        # Move to next heat
        heat_index += direction

        # Reverse direction at ends (snake pattern)
        if heat_index >= num_heats:
            heat_index = num_heats - 1
            direction = -1
        elif heat_index < 0:
            heat_index = 0
            direction = 1

    # Build competitor DataFrames and set advancement rules for each heat
    for heat in heats:
        # Get DataFrame subset for competitors in this heat
        heat['competitors_df'] = all_competitors_df[
            all_competitors_df['competitor_name'].isin(heat['competitors'])
        ].copy()

        # Determine advancement rules based on heat size
        heat_size = len(heat['competitors'])
        heat_capacity = num_stands
        fill_percentage = heat_size / heat_capacity

        # If heat is ≤50% full, only top 1 advances (fairness for partial heats)
        if fill_percentage <= 0.5:
            heat['num_to_advance'] = 1
        else:
            heat['num_to_advance'] = 2

    return heats


def select_heat_advancers(round_object):
    """Interactive function for judge to select advancing competitors from a completed heat.

    Args:
        round_object (dict): Round object for completed heat

    Returns:
        list: Names of advancing competitors
    """

    print(f"\n{'='*70}")
    print(f"  {round_object['round_name'].upper()} RESULTS & ADVANCERS")
    print(f"{'='*70}")

    # Use FINISH ORDER to determine results (critical for handicap racing)
    if round_object.get('finish_order'):
        # Sort by finish order (1st, 2nd, 3rd, etc.)
        sorted_by_finish = sorted(
            round_object['finish_order'].items(),
            key=lambda x: x[1]  # Sort by finish position
        )

        # Display results table with finish position, cutting time, and advancement status
        print("\n╔════════════════════════════════════════════════════════════════════╗")
        print(f"║           🏆 {round_object['round_name'].upper()} - FINAL RESULTS 🏆            ║")
        print("╠════════════════════════════════════════════════════════════════════╣")
        print("║  Finish Position based on REAL-TIME completion (handicap included) ║")
        print("╠════════════════════════════════════════════════════════════════════╣")

        advancers = []  # Auto-select based on finish order

        for name, finish_pos in sorted_by_finish:
            cutting_time = round_object['actual_results'].get(name, 0)

            # Determine medal and advancement
            medal = ""
            advance = ""
            if finish_pos == 1:
                medal = "🥇"
            elif finish_pos == 2:
                medal = "🥈"
            elif finish_pos == 3:
                medal = "🥉"
            else:
                medal = "  "

            # Auto-select top finishers for advancement
            if finish_pos <= round_object['num_to_advance']:
                advance = "✓ ADVANCES"
                advancers.append(name)

            # Format line: Position | Name | Cutting Time | Status
            pos_part = f"{medal} {finish_pos:2d}"
            name_part = f"{name[:35]:<35}"
            time_part = f"{cutting_time:6.2f}s"
            status_part = f"{advance:^12}"

            print(f"║  {pos_part} │ {name_part} │ {time_part} │ {status_part} ║")

        print("╚════════════════════════════════════════════════════════════════════╝\n")

        # Display advancers summary
        print(f"✓ Top {round_object['num_to_advance']} finisher(s) automatically advance:")
        for name in advancers:
            finish_pos = round_object['finish_order'][name]
            print(f"  {finish_pos}. {name}")

        # Allow judge to override if needed
        override = input("\nAccept these advancers? (y/n to manually select): ").strip().lower()

        if override == 'y' or override == '':
            # Accept auto-selected advancers
            round_object['advancers'] = advancers
            round_object['status'] = 'completed'
            return advancers
        else:
            # Manual override - let judge pick
            print("\nManual selection mode:")
            advancers = []
            while len(advancers) < round_object['num_to_advance']:
                try:
                    for i, (name, pos) in enumerate(sorted_by_finish, 1):
                        print(f"  {i}) {name} (finished {pos})")

                    choice = input(f"\nSelect competitor {len(advancers)+1} of {round_object['num_to_advance']} (number): ").strip()
                    idx = int(choice) - 1

                    if 0 <= idx < len(sorted_by_finish):
                        selected = sorted_by_finish[idx][0]
                        if selected not in advancers:
                            advancers.append(selected)
                            print(f"  ✓ {selected} selected")
                        else:
                            print("  Already selected. Choose another competitor.")
                    else:
                        print(f"  Invalid number. Enter 1-{len(sorted_by_finish)}")

                except ValueError:
                    print("  Please enter a valid number.")

    else:
        # Fallback: No finish order recorded (legacy mode - use raw times)
        print("⚠ WARNING: No finish order recorded. Using raw cutting times (may be inaccurate for handicap).\n")

        if round_object['actual_results']:
            sorted_results = sorted(
                round_object['actual_results'].items(),
                key=lambda x: x[1]  # Sort by raw time
            )

            print("Competitors (sorted by cutting time):")
            for i, (name, time) in enumerate(sorted_results, 1):
                print(f"  {i}) {name:30s} {time:.2f}s")
        else:
            print("Competitors in heat:")
            for i, name in enumerate(round_object['competitors'], 1):
                print(f"  {i}) {name}")

        # Manual selection
        advancers = []
        while len(advancers) < round_object['num_to_advance']:
            try:
                choice = input(f"\nSelect competitor {len(advancers)+1} of {round_object['num_to_advance']} (number): ").strip()

                if choice == '':
                    print("Selection cannot be blank. Please enter a number.")
                    continue

                idx = int(choice) - 1

                if 0 <= idx < len(round_object['competitors']):
                    selected = round_object['competitors'][idx]
                    if selected not in advancers:
                        advancers.append(selected)
                        print(f"  ✓ {selected} selected")
                    else:
                        print("  Already selected. Choose another competitor.")
                else:
                    print(f"  Invalid number. Enter 1-{len(round_object['competitors'])}")

            except ValueError:
                print("  Please enter a valid number.")

    # Update round object
    round_object['advancers'] = advancers
    round_object['status'] = 'completed'

    return advancers


def generate_next_round(tournament_state, all_advancers, next_round_type):
    """Generate semi-final or final rounds from advancing competitors.

    Args:
        tournament_state (dict): Global tournament state
        all_advancers (list): All competitors advancing from previous round
        next_round_type (str): 'semi' or 'final'

    Returns:
        list: List of round_object dictionaries for next stage
    """

    # Extract handicap results for advancers only
    all_results = []
    for round_obj in tournament_state['rounds']:
        if 'handicap_results' in round_obj and round_obj['handicap_results']:
            all_results.extend(round_obj['handicap_results'])

    # Filter to just advancers
    advancer_results = [r for r in all_results if r['name'] in all_advancers]

    # Determine number of heats for next round
    num_stands = tournament_state['num_stands']

    if next_round_type == 'final':
        # Single final heat
        num_heats = 1
    elif next_round_type == 'semi':
        # Multiple semi-finals based on advancer count
        num_heats = ceil(len(all_advancers) / num_stands)
    else:
        num_heats = ceil(len(all_advancers) / num_stands)

    # Get DataFrame for advancers only
    all_advancers_df = tournament_state['all_competitors_df'][
        tournament_state['all_competitors_df']['competitor_name'].isin(all_advancers)
    ].copy()

    # Use same distribution algorithm (snake draft)
    next_rounds = distribute_competitors_into_heats(
        all_advancers_df,
        advancer_results,
        num_stands,
        num_heats
    )

    # Update round type and names
    for i, round_obj in enumerate(next_rounds):
        round_obj['round_type'] = next_round_type
        if next_round_type == 'final':
            round_obj['round_name'] = 'Final'
            round_obj['round_number'] = 1
        elif next_round_type == 'semi':
            round_obj['round_name'] = f'Semi {i + 1}'
            round_obj['round_number'] = i + 1

    return next_rounds


def view_tournament_status(tournament_state):
    """Display visual tournament bracket showing progress and results.

    Args:
        tournament_state (dict): Global tournament state
    """

    print(f"\n{'='*70}")
    print(f"  TOURNAMENT STATUS")
    print(f"{'='*70}")

    if not tournament_state.get('rounds'):
        print("\nNo tournament rounds generated yet.")
        return

    print(f"\nEvent: {tournament_state.get('event_name', 'Unnamed Event')}")
    print(f"Format: {tournament_state.get('format', 'Unknown')}")
    print(f"Stands: {tournament_state.get('num_stands', '?')}")
    print(f"Total Competitors: {len(tournament_state.get('all_competitors', []))}")

    # Group rounds by type
    heats = [r for r in tournament_state['rounds'] if r['round_type'] == 'heat']
    semis = [r for r in tournament_state['rounds'] if r['round_type'] == 'semi']
    finals = [r for r in tournament_state['rounds'] if r['round_type'] == 'final']

    # Display heats
    if heats:
        print(f"\n{'─'*70}")
        print(f"INITIAL HEATS ({len(heats)} total)")
        print(f"{'─'*70}")
        for heat in heats:
            status_icon = "✓" if heat['status'] == 'completed' else "○"
            print(f"{status_icon} {heat['round_name']:15s} - {len(heat['competitors'])} competitors, top {heat['num_to_advance']} advance")
            if heat['status'] == 'completed' and heat.get('advancers'):
                print(f"    Advancers: {', '.join(heat['advancers'])}")

    # Display semis
    if semis:
        print(f"\n{'─'*70}")
        print(f"SEMI-FINALS ({len(semis)} total)")
        print(f"{'─'*70}")
        for semi in semis:
            status_icon = "✓" if semi['status'] == 'completed' else "○"
            print(f"{status_icon} {semi['round_name']:15s} - {len(semi['competitors'])} competitors, top {semi['num_to_advance']} advance")
            if semi['status'] == 'completed' and semi.get('advancers'):
                print(f"    Advancers: {', '.join(semi['advancers'])}")

    # Display finals
    if finals:
        print(f"\n{'─'*70}")
        print(f"FINAL")
        print(f"{'─'*70}")
        for final in finals:
            status_icon = "✓" if final['status'] == 'completed' else "○"
            print(f"{status_icon} {final['round_name']:15s} - {len(final['competitors'])} competitors")
            if final['status'] == 'completed':
                print(f"    Tournament Complete!")

    print(f"\n{'='*70}\n")


def save_tournament_state(tournament_state, filename="tournament_state.json"):
    """Save tournament state to JSON for crash recovery.

    Args:
        tournament_state (dict): Tournament state to save
        filename (str): Output filename
    """
    import json

    try:
        # Convert DataFrames to dict format for JSON serialization
        state_copy = tournament_state.copy()

        # Convert main competitors DataFrame
        if not state_copy['all_competitors_df'].empty:
            state_copy['all_competitors_df'] = state_copy['all_competitors_df'].to_dict('records')
        else:
            state_copy['all_competitors_df'] = []

        # Convert DataFrames in rounds
        for round_obj in state_copy.get('rounds', []):
            if not round_obj['competitors_df'].empty:
                round_obj['competitors_df'] = round_obj['competitors_df'].to_dict('records')
            else:
                round_obj['competitors_df'] = []

        # Write to file
        with open(filename, 'w') as f:
            json.dump(state_copy, f, indent=2, default=str)

        print(f"Tournament state saved to {filename}")

    except Exception as e:
        print(f"Error saving tournament state: {e}")


def load_tournament_state(filename="tournament_state.json"):
    """Load tournament state from JSON.

    Args:
        filename (str): Input filename

    Returns:
        dict: Loaded tournament state, or None if error
    """
    import json

    try:
        with open(filename, 'r') as f:
            state = json.load(f)

        # Convert dict records back to DataFrames
        if state.get('all_competitors_df'):
            state['all_competitors_df'] = pd.DataFrame(state['all_competitors_df'])
        else:
            state['all_competitors_df'] = pd.DataFrame()

        # Convert DataFrames in rounds
        for round_obj in state.get('rounds', []):
            if round_obj.get('competitors_df'):
                round_obj['competitors_df'] = pd.DataFrame(round_obj['competitors_df'])
            else:
                round_obj['competitors_df'] = pd.DataFrame()

        print(f"Tournament state loaded from {filename}")
        return state

    except FileNotFoundError:
        print(f"Tournament state file '{filename}' not found.")
        return None
    except Exception as e:
        print(f"Error loading tournament state: {e}")
        return None


def auto_save_state(tournament_state):
    """Auto-save tournament state after significant actions.

    Args:
        tournament_state (dict): Tournament state to save
    """
    save_tournament_state(tournament_state, "tournament_state.json")


# ============================================================================
# END OF TOURNAMENT MANAGEMENT FUNCTIONS
# ============================================================================


def personnel_management_menu(comp_df):
    """Personnel Management Menu - for adding/editing/removing competitors from master roster.

    This is separated from tournament operations to allow roster management
    before or during tournament setup.

    Args:
        comp_df (DataFrame): Current competitor roster

    Returns:
        DataFrame: Updated competitor roster
    """

    while True:
        print("\n" + "=" * 60)
        print("  PERSONNEL MANAGEMENT MENU")
        print("=" * 60)
        print("1) Add New Competitor to Roster")
        print("2) View Full Roster")
        print("3) Remove Competitor from Roster")
        print("4) Back to Main Menu")
        print("=" * 60)

        choice = input("Choose an option: ").strip()

        if choice == "1":
            # Add new competitor with historical times
            comp_df = add_competitor_with_times()
            print("\n✓ Competitor added successfully")

        elif choice == "2":
            # View full roster
            if comp_df is None or comp_df.empty:
                print("\nRoster is empty. No competitors found.")
            else:
                print(f"\n{'='*60}")
                print(f"  FULL ROSTER ({len(comp_df)} competitors)")
                print(f"{'='*60}")
                for idx in range(len(comp_df)):
                    row = comp_df.iloc[idx]
                    name = row.get("competitor_name", "Unknown")
                    country = row.get("competitor_country", "Unknown")
                    comp_id = row.get("competitor_id", "N/A")
                    print(f"{idx + 1:3d}) {name:30s} ({country:15s}) [ID: {comp_id}]")
                print(f"{'='*60}")

        elif choice == "3":
            # Remove competitor from roster
            print("\nRemove Competitor - Feature Coming Soon")
            print("(Currently, please remove directly from Excel file)")

        elif choice == "4" or choice == "":
            break

        else:
            print("Invalid selection. Try again.")

    return comp_df


def select_all_event_competitors(comp_df, max_competitors=None):
    """Select ALL competitors for a tournament event (not just one heat).

    This replaces the old select_competitors_for_heat() function for tournament mode.
    Supports multi-select via comma-separated or range input.

    Args:
        comp_df (DataFrame): Full competitor roster
        max_competitors (int): Maximum allowed competitors (optional)

    Returns:
        DataFrame: Selected competitors
    """

    if comp_df is None or comp_df.empty:
        print("\nRoster is empty. Please add competitors first.")
        input("Press Enter to continue...")
        return pd.DataFrame()

    if "competitor_name" not in comp_df.columns:
        print("Roster missing 'competitor_name' column.")
        input("Press Enter to continue...")
        return pd.DataFrame()

    # Display roster with index numbers
    print(f"\n{'='*70}")
    print(f"  SELECT COMPETITORS FOR EVENT")
    print(f"{'='*70}")

    if max_competitors:
        print(f"Maximum competitors for this format: {max_competitors}\n")

    print("ROSTER (All Available Competitors):")
    print("-" * 70)
    for idx in range(len(comp_df)):
        row = comp_df.iloc[idx]
        name = row.get("competitor_name", "Unknown")
        country = row.get("competitor_country", "Unknown")
        print(f"  {idx + 1:3d}) {name:35s} ({country})")

    print("\n" + "=" * 70)
    print("INSTRUCTIONS:")
    print("  - Enter single number: 5")
    print("  - Enter multiple numbers (comma-separated): 1,3,5,7")
    print("  - Enter range: 1-10")
    print("  - Combine: 1,3,5-8,12")
    print("  - Press Enter with no input when finished")
    print("=" * 70)

    selected_indices = set()

    while True:
        selection = input(f"\nEnter competitor number(s) (or press Enter to finish): ").strip()

        if selection == "":
            break

        # Parse input (supports ranges and comma-separated)
        try:
            for part in selection.split(','):
                part = part.strip()
                if '-' in part:
                    # Range: e.g., "5-10"
                    start, end = part.split('-')
                    start_idx = int(start) - 1
                    end_idx = int(end) - 1
                    for i in range(start_idx, end_idx + 1):
                        if 0 <= i < len(comp_df):
                            selected_indices.add(i)
                        else:
                            print(f"  ⚠ Skipping invalid number: {i+1}")
                else:
                    # Single number
                    idx = int(part) - 1
                    if 0 <= idx < len(comp_df):
                        selected_indices.add(idx)
                    else:
                        print(f"  ⚠ Invalid number: {part}")

            # Show current selection count
            print(f"  ✓ {len(selected_indices)} competitor(s) selected")

            # Check max limit
            if max_competitors and len(selected_indices) > max_competitors:
                print(f"  ⚠ WARNING: {len(selected_indices)} exceeds maximum of {max_competitors}")
                over = input(f"    Continue anyway? (y/n): ").strip().lower()
                if over != 'y':
                    print("  Resetting selection...")
                    selected_indices = set()

        except ValueError:
            print("  ⚠ Invalid input. Use format: 1,3,5-8,12")

    if not selected_indices:
        print("\nNo competitors selected.")
        return pd.DataFrame()

    # Build DataFrame of selected competitors
    selected_df = comp_df.iloc[sorted(selected_indices)].copy()

    # Display final selection
    print(f"\n{'='*70}")
    print(f"  FINAL SELECTION ({len(selected_df)} competitors)")
    print(f"{'='*70}")
    for idx, row in enumerate(selected_df.iterrows(), 1):
        _, data = row
        name = data.get("competitor_name", "Unknown")
        country = data.get("competitor_country", "Unknown")
        print(f"  {idx:3d}) {name:35s} ({country})")
    print(f"{'='*70}")

    confirm = input("\nConfirm this selection? (y/n): ").strip().lower()
    if confirm != 'y':
        print("Selection cancelled. Returning to menu.")
        return pd.DataFrame()

    return selected_df


# ============================================================================
# END OF NEW TOURNAMENT SUPPORT FUNCTIONS
# ============================================================================


#Menu Option 1: Competitor Selection Menu
''' Official will be presented with a list of competitors

    Definitions:
    'Roster'- list of all competitors available in the excel sheet
    'Heat Assignment'- list of competitors selected for the current heat

    1. Select Competitors for Heat from the roster
        a. Reload roster from Excel to ensure we have latest data
        b. Competitors will be displayed with index numbers for selection
        C. Selected competitors will be added to heat assignment list
    2. Add competitors to the roster
        a. Prompt for competitor details (name, country)
        b. Input historcical times for handicap calculation (3x)
            -Event (UH/SB)
            -Time
            -Wood species
            -Size in mm
            -Date (optional)
    3. View Heat Assignment
    -While this is functionally similar to viewing the heat after selecting competitors,
    this allows the judge to review the heat at any time.
    4. Remove competitors from the heat assignment
    -Self Explanatory; this does NOT remove them from the roster, only from the current heat assignment
    5. Return to Main Menu
'''

## Competitor Selection Menu
def competitor_menu(comp_df, heat_assignment_df, heat_assignment_names):

    while True:
        print("\n--- Competitor Menu ---")
        print("1) Select Competitors for Heat from roster")
        print("2) Add new competitor to roster")
        print("3) View Heat Assignment")
        print("4) Remove competitor from Heat Assignment")
        print("5) Back to Main Menu")
        s = input("Choose an option: ").strip()

        if s == "1":
            # Reload roster from Excel to ensure we have latest data
            comp_df = load_competitors_df()
            
            # Select competitors and RETURN TO MAIN MENU
            heat_assignment_df, heat_assignment_names = select_competitors_for_heat(comp_df)
            return comp_df, heat_assignment_df, heat_assignment_names

        elif s == "2":
            # Add new competitor to roster with historical times
            comp_df = add_competitor_with_times()
            # Stay in competitor menu after adding

        elif s == "3":
            # View current heat assignment
            view_heat_assignment(heat_assignment_df, heat_assignment_names)
            # Stay in competitor menu after viewing

        elif s == "4":
            # Remove from heat assignment (not from roster)
            heat_assignment_df, heat_assignment_names = remove_from_heat(heat_assignment_df, heat_assignment_names)
            # Stay in competitor menu after removing

        elif s == "5" or s == "":
            break
        else:
            print("Invalid selection. Try again.")
    
    return comp_df, heat_assignment_df, heat_assignment_names


def select_competitors_for_heat(comp_df):

    """Select competitors from roster to add to heat assignment will display the names of all competitors available in the excel sheet.
    -All competitors will have an index number (different from the competitor ID that is on the excel shet!) assigned to them for easy selection
    -The Judge will enter a competitor's index number to select them for the heat
    -Selected competitors will be added to a separate list for the heat
    -Slected competitors' names will be displayed after selection is complete
    -The judge will be prompted to hit enter on an empty entry to finalize list
"""
    #Obligatory checks

    if comp_df is None or comp_df.empty:
        print("\nRoster is empty. Please add competitors first.")
        input("Press Enter to continue...")
        return pd.DataFrame(), []
    
    if "competitor_name" not in comp_df.columns:
        print("Roster missing 'competitor_name' column.")
        input("Press Enter to continue...")
        return pd.DataFrame(), []
    
    # Display roster with index numbers
    print("\n--- ROSTER (All Available Competitors) ---")
    print("-" * 40)
    for idx in range(len(comp_df)):
        row = comp_df.iloc[idx]
        name = row.get("competitor_name", "Unknown")
        country = row.get("competitor_country", "Unknown")
        print(f"{idx + 1}) {name} ({country})")
    
    print("\n" + "=" * 40)
    print("INSTRUCTIONS:")
    print("- Enter competitor numbers one at a time")
    print("- Press Enter after each number to add them")
    print("- Press Enter with no input when finished")
    print("=" * 40)
    
    selected_indices = []
    selected_names = []
    
    while True:
        selection = input(f"\nEnter competitor number (or press Enter to finish): ").strip()
        
        if selection == "":
            break
        
        try:
            idx = int(selection) - 1
            if 0 <= idx < len(comp_df):
                if idx not in selected_indices:
                    selected_indices.append(idx)
                    name = comp_df.iloc[idx]["competitor_name"]
                    selected_names.append(name)
                    print(f"✓ {name} added to heat")
                else:
                    print("Competitor already selected")
            else:
                print("Invalid number. Please try again.")
        except ValueError:
            print("Please enter a valid number")
    
    if not selected_indices:
        print("\nNo competitors selected.")
        input("Press Enter to return to main menu...")
        return pd.DataFrame(), []
    
    heat_df = comp_df.iloc[selected_indices].copy()
    
    print(f"\n✓ Total {len(selected_names)} competitors added to heat assignment:")
    for name in selected_names:
        print(f"  - {name}")
    
    input("\nPress Enter to return to main menu...")
    return heat_df, selected_names

# Display current heat assignment. This will display the names of all competitors currently selected for the heat
    

def view_heat_assignment(heat_df, heat_names):
    print("\n--- CURRENT HEAT ASSIGNMENT ---")
    print("-" * 40)
    
    if not heat_names or heat_df.empty:
        print("No competitors currently assigned to heat.")
    else:
        for i, name in enumerate(heat_names, 1):
            country = heat_df[heat_df["competitor_name"] == name]["competitor_country"].values
            country_str = country[0] if len(country) > 0 else "Unknown"
            print(f"{i}) {name} ({country_str})")
        
        print(f"\nTotal: {len(heat_names)} competitors in heat")
    
    input("\nPress Enter to continue...")

# Remove competitor from heat assignment (not from roster). This will allow the judge to remove competitors from the heat assignment list
def remove_from_heat(heat_df, heat_names):
  
    if not heat_names or heat_df.empty:
        print("\nHeat assignment is currently empty.")
        input("Press Enter to continue...")
        return heat_df, heat_names
    
    print("\n--- REMOVE FROM HEAT ASSIGNMENT ---")
    print("-" * 40)
    
    for i, name in enumerate(heat_names, 1):
        print(f"{i}) {name}")
    
    print("\nEnter number to remove or press Enter to cancel:")
    choice = input("Your choice: ").strip()
    
    if choice == "":
        print("No changes made.")
        input("Press Enter to continue...")
        return heat_df, heat_names
    
    try:
        idx = int(choice) - 1
        if 0 <= idx < len(heat_names):
            removed_name = heat_names[idx]
            heat_names.remove(removed_name)
            heat_df = heat_df[heat_df["competitor_name"] != removed_name]
            print(f"\n✓ {removed_name} removed from heat assignment.")
        else:
            print("Invalid selection.")
    except ValueError:
        print("Invalid input.")
    
    input("Press Enter to continue...")
    return heat_df, heat_names


def add_competitor_with_times():
    """Add a new competitor to roster and prompt for historical times"""
    try:
        # Get competitor basic info
        print("\n--- Add New Competitor ---")
        name = input("Enter competitor name: ").strip()
        if not name:
            print("Name cannot be empty.")
            return load_competitors_df()
        
        country = input("Enter competitor country: ").strip()
        state = input("Enter state/province (optional): ").strip()
        gender = input("Enter gender (M/F, optional): ").strip().upper()
        
        # Add to competitors sheet
        try:
            wb = load_workbook(COMPETITOR_FILE)
        except FileNotFoundError:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
        
        if COMPETITOR_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(COMPETITOR_SHEET)
            ws.append(["CompetitorID", "Name", "Country", "State/Province", "Gender"])
        else:
            ws = wb[COMPETITOR_SHEET]
            if ws.max_row < 1:
                ws.append(["CompetitorID", "Name", "Country", "State/Province", "Gender"])
        
        # Check for duplicate
        existing_names = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) > 1 and row[1]:  # Name is in column 2 (index 1)
                existing_names.add(str(row[1]).strip().lower())
        
        if name.lower() in existing_names:
            print("Competitor already exists in roster.")
            wb.close()
            return load_competitors_df()
        
        # Generate new CompetitorID
        new_id = f"C{str(ws.max_row).zfill(3)}"
        
        # Add competitor to sheet
        ws.append([new_id, name, country, state, gender])
        
        wb.save(COMPETITOR_FILE)
        wb.close()
        print(f"\n✓ {name} added to roster successfully with ID {new_id}")
        
        # Now prompt for historical times
        print("\n--- Enter Historical Competition Times ---")
        print("Minimum 3 times required for handicap calculation.")
        print("You can enter more than 3 if desired.")
        
        add_historical_times_for_competitor(name)
        
        return load_competitors_df()
        
    except Exception as e:
        print(f"Error adding competitor: {e}")
        return load_competitors_df()
    

def add_historical_times_for_competitor(competitor_name):
    """Prompt for and save historical times to results sheet
    Judge will be prompted to select whether each time is for SB or UH
    Judge will be prompted to enter the time, wood species, size, and date (optional)
    The program will store this data in the results sheet
    """
    times_added = 0
    
    while True:
        if times_added < 3:
            print(f"\n--- Historical Time Entry {times_added + 1} (minimum 3 required) ---")
        else:
            cont = input(f"\n{times_added} times entered. Add another? (y/n): ").strip().lower()
            if cont != 'y':
                break
            print(f"\n--- Historical Time Entry {times_added + 1} ---")
        
        # Get event type
        while True:
            event = input("Event type (SB for Standing Block, UH for Underhand): ").strip().upper()
            if event in ["SB", "UH"]:
                break
            print("Please enter SB or UH.")
        
        # Get time
        while True:
            time_str = input("Time in seconds (e.g., 45.3): ").strip()
            try:
                time_val = float(time_str)
                break
            except ValueError:
                print("Please enter a valid number.")
        
        # Get wood species
        species = input("Wood species: ").strip()
        
        # Get size
        while True:
            size_str = input("Wood diameter in mm: ").strip()
            try:
                size_val = float(size_str)
                break
            except ValueError:
                print("Please enter a valid number.")
        
        # Get quality (optional, default to 5)
        quality_str = input("Wood quality (0-10, press Enter for 5): ").strip()
        if quality_str:
            try:
                quality = int(quality_str)
                quality = max(0, min(10, quality))
            except:
                quality = 5
        else:
            quality = 5
        
        # Optional date
        date_str = input("Date (optional, format YYYY-MM-DD or press Enter to skip): ").strip()
        if not date_str:
            timestamp = datetime.now().isoformat(timespec="seconds")
        else:
            try:
                date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                timestamp = date_obj.isoformat(timespec="seconds")
            except:
                timestamp = datetime.now().isoformat(timespec="seconds")
        
        # Save to results sheet
        save_time_to_results(event, competitor_name, species, size_val, quality, time_val, 
                           f"Historical-{event}", timestamp)
        times_added += 1
        print(f"✓ Time #{times_added} saved successfully")
    
    if times_added >= 3:
        print(f"\n✓ {times_added} historical times added for {competitor_name}.")
    else:
        print(f"\n⚠ Warning: Only {times_added} times added. Minimum 3 recommended for handicap calculation.")


def save_time_to_results(event, name, species, size, quality, time, heat_id, timestamp):
    """Helper to save a single time entry to results sheet"""
    try:
        # Convert name to ID
        _, name_to_id = get_competitor_id_name_mapping()
        competitor_id = name_to_id.get(str(name).strip().lower(), name)
        
        try:
            wb = load_workbook(RESULTS_FILE)
        except:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
        
        if RESULTS_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(RESULTS_SHEET)
            ws.append(["CompetitorID", "Event", "Time (seconds)", "Size (mm)", "Species Code", "Date"])
        else:
            ws = wb[RESULTS_SHEET]
            if ws.max_row == 0:
                ws.append(["CompetitorID", "Event", "Time (seconds)", "Size (mm)", "Species Code", "Date"])
        
        # Save with Excel's expected column names
        ws.append([competitor_id, event, time, size, species, timestamp])
        wb.save(RESULTS_FILE)
        wb.close()
        
    except Exception as e:
        print(f"Error saving time to results: {e}")



# MENU OPTION 2: WOOD CHARACTERISTICS MENU

''' Official will be presented with a list of wood characteristics:
    1. Select Wood Species from the list
    2. Enter Size in mm
    3. Enter wood quality 
    4. Return to Main Menu

    This menu will allow the judge to select the characteristics of the wood block being used in the heat and store the 
    selection for handicap calculation.
    Wood species available will be loaded from the wood sheet in the excel file.
'''

## Wood Characteristics Menu 
def wood_menu(wood_selection):
    
    if "event" not in wood_selection:
        wood_selection["event"] = None

    while True:
        print("\n--- Wood Menu ---")
        print(f"Current: species={wood_selection.get('species')}, "
              f"size_mm={wood_selection.get('size_mm')}, "
              f"quality={wood_selection.get('quality')}")
        print("1) Select wood species")
        print("2) Enter size (mm)")
        print("3) Enter quality (0 for poor quality, 1-3 for soft wood, 4-7 for average firmness for species, 8-10 for above average firmness for species)")
        print("4) Back to Main Menu")
        s = input("Choose an option: ").strip()

        if s == "1":
            wood_selection = select_wood_species(wood_selection)

        elif s == "2":
            wood_selection = enter_wood_size_mm(wood_selection)

        elif s == "3":
            wood_selection = enter_wood_quality(wood_selection)

        elif s == "4" or s == "":
            break
        else:
            print("Invalid selection. Try again.")
    
    return wood_selection


##Display species list, accept numeric choice
def select_wood_species(wood_selection):
    """Select wood species"""
    wood_df = load_wood_data()
    
    if wood_df.empty:
        print("No wood data available.")
        return wood_selection
    
    if "species" not in wood_df.columns:
        print("Wood sheet missing 'species' column.")
        return wood_selection
    
    species_list = wood_df["species"].astype(str).tolist()
    
    print("\nAvailable wood species:")
    for i, sp in enumerate(species_list, start=1):
        print(f"{i}) {sp}")
    
    choice = input("Select species by number: ").strip()
    
    try:
        idx = int(choice) - 1
        if 0 <= idx < len(species_list):
            wood_selection["species"] = species_list[idx]
            format_wood(wood_selection)
        else:
            print("Invalid selection.")
    except ValueError:
        print("Invalid input.")
    
    return wood_selection


## Enter block size in mm
def enter_wood_size_mm(wood_selection):
    """Enter block size in mm"""
    size = input("Enter block diameter in mm: ").strip()
    
    try:
        val = float(size)
        wood_selection["size_mm"] = val
        format_wood(wood_selection)
    except ValueError:
        print("Invalid size input.")
    
    return wood_selection


##Prompt for quality (integer 0–10)
'''Higher quality means softer wood and a faster time. 
this needs to account for that because softer wood would favor the front/back marker differently. 
A "0" indicates wood barely suitable for competition.'''
def enter_wood_quality(wood_selection):
    """Enter wood quality rating as an integer 0–10"""
    while True:
        s = input("Enter wood quality (integer 0–10): ").strip()
        
        if s == "":
            print("No change made to wood quality.")
            break
        
        try:
            val = int(s)
            val = max(0, min(10, val))  # Clamp between 0 and 10
            wood_selection["quality"] = val
            format_wood(wood_selection)
            break
        except ValueError:
            print("Invalid input. Please enter an integer between 0 and 10.")
    
    return wood_selection


##Header that displays current wood selection
def format_wood(ws):
    """Return formatted header for wood selection"""
    s = ws.get("species", "—")
    d = ws.get("size_mm", "—")
    q = ws.get("quality", "—")
    header = f"Selected Wood -> Species: {s}, Diameter: {d} mm, Quality: {q}"
    print(f"Wood selection updated: {header}")
    return header


# MENU OPTION 3: SELECT EVENT (SB/UH)
''' Official will be presented with two event codes to select from either SB or UH
Prompt for event code and store as SB or UH. No other values allowed.'''

def select_event_code(wood_selection):
    while True:
        e = input("Select event: type 'SB' for Standing Block or 'UH' for Underhand: ").strip().upper()
        
        if e in ("SB", "UH"):
            wood_selection["event"] = e
            print(f"Event selected: {e}")
            return wood_selection
        
        if e == "":
            print("No change made to event.")
            return wood_selection
        
        print("Invalid input. Please enter SB or UH.")


# AI INTEGRATION - OLLAMA FUNCTIONS (USED BY MENU 4)

#Call in OLLAMA (qwen2.5:7b)
def call_ollama(prompt: str, model: str = None) -> Optional[str]:
    """
    Send prompt to local Ollama instance and return response.

    Args:
        prompt: Text prompt to send to the model
        model: Ollama model name (defaults to config value)

    Returns:
        Model response text, or None if error occurs
    """
    if model is None:
        model = llm_config.DEFAULT_MODEL

    try:
        response = requests.post(
            llm_config.OLLAMA_URL,
            json={
                "model": model,
                "prompt": prompt,
                "stream": False,
                "options": {
                    "temperature": 0.3,  # temperature: 0.3 = Low creativity.
                    "num_predict": 50    # Limit response length for speed
                }
            },
            timeout=llm_config.TIMEOUT_SECONDS
        )
        
        if response.status_code == 200:
            return response.json()['response'].strip()
        else:
            print(f"Ollama error: {response.status_code}")
            return None
            
    except requests.exceptions.ConnectionError:
        print("\nError: Cannot connect to Ollama. Make sure it's running:")
        print("  Run 'ollama serve' in a terminal")
        return None
    except Exception as e:
        print(f"\nError calling Ollama: {e}")
        return None




def get_competitor_historical_times_flexible(competitor_name, species, event_code, results_df):
  
    if results_df is None or results_df.empty:
        return [], "no data available"
    
    # Match competitor and event (required)
    name_match = results_df["competitor_name"].astype(str).str.strip().str.lower() == str(competitor_name).strip().lower()
    event_match = results_df["event"].astype(str).str.strip().str.upper() == event_code
    
    # Try exact species match first
    if species and "species" in results_df.columns:
        species_match = results_df["species"].astype(str).str.strip().str.lower() == str(species).strip().lower()
        exact_matches = results_df[name_match & event_match & species_match]
        
        times = []
        for _, row in exact_matches.iterrows():
            time = row.get('raw_time')
            if time is not None and time > 0:
                times.append(time)
        
        if times:
            return times, f"on {species} (exact match)"
    
    # Fallback: any species for this competitor and event
    any_species_matches = results_df[name_match & event_match]
    times = []
    for _, row in any_species_matches.iterrows():
        time = row.get('raw_time')
        if time is not None and time > 0:
            times.append(time)
    
    if times:
        return times, "on various wood types"
    
    return [], "no competitor history found"


#Calcualte Baseline

'''Calculate baseline with cascading fallback.

Tries in order:
1. Species + diameter range + event
2. Diameter range + event (any species)
3. Event only (any species, any diameter)
'''
    

def get_event_baseline_flexible(species, diameter, event_code, results_df):
  
    if results_df is None or results_df.empty:
        return None, "no data available"
    
    event_match = results_df["event"].astype(str).str.strip().str.upper() == event_code
    
    # Try species + diameter range + event
    if species and "species" in results_df.columns and "size_mm" in results_df.columns:
        species_match = results_df["species"].astype(str).str.strip().str.lower() == str(species).strip().lower()
        diameter_match = (results_df["size_mm"] >= diameter - 25) & (results_df["size_mm"] <= diameter + 25)
        
        exact_matches = results_df[species_match & diameter_match & event_match]
        times = []
        for _, row in exact_matches.iterrows():
            time = row.get('raw_time')
            if time is not None and time > 0:
                times.append(time)
        
        if len(times) >= 3:
            return statistics.mean(times), f"species/size average ({len(times)} performances)"
    
    # Fallback: diameter range + event (any species)
    if "size_mm" in results_df.columns:
        diameter_match = (results_df["size_mm"] >= diameter - 25) & (results_df["size_mm"] <= diameter + 25)
        size_matches = results_df[diameter_match & event_match]
        
        times = []
        for _, row in size_matches.iterrows():
            time = row.get('raw_time')
            if time is not None and time > 0:
                times.append(time)
        
        if len(times) >= 3:
            return statistics.mean(times), f"size average ({len(times)} performances)"
    
    # Final fallback: event only (all data for this event type)
    event_only = results_df[event_match]
    times = []
    for _, row in event_only.iterrows():
        time = row.get('raw_time')
        if time is not None and time > 0:
            times.append(time)
    
    if len(times) >= 3:
        return statistics.mean(times), f"event average ({len(times)} performances)"
    
    return None, "insufficient data"


#AI predicted Handicaps:

    
''' Predict competitor's time using historical data + LLM reasoning for quality adjustment.
    Now with improved fallback logic for sparse data.
    
'''

def predict_competitor_time_with_ai(competitor_name, species, diameter, quality, event_code, results_df):
    ''' Predict competitor's time using historical data + LLM reasoning for quality adjustment.'''
   
    # Step 1: Get historical data
    historical_times, data_source = get_competitor_historical_times_flexible(
        competitor_name, species, event_code, results_df
    )
    
    # Step 2: Calculate baseline
    if len(historical_times) >= 3:
        weights = [1.5, 1.3, 1.1, 1.0, 0.9, 0.8]
        weighted_times = [t * w for t, w in zip(historical_times[:6], weights[:len(historical_times)])]
        baseline = sum(weighted_times) / sum(weights[:len(historical_times)])
        confidence = "HIGH"
        explanation_source = f"competitor history {data_source}"
        
    elif len(historical_times) > 0:
        baseline = statistics.mean(historical_times)
        confidence = "MEDIUM"
        explanation_source = f"limited competitor history {data_source}"
        
    else:
        baseline, baseline_source = get_event_baseline_flexible(species, diameter, event_code, results_df)
        
        if baseline:
            confidence = "LOW"
            explanation_source = f"{baseline_source} (no competitor history)"
        else:
            if diameter >= 350:
                baseline = 60.0
            elif diameter >= 300:
                baseline = 50.0
            elif diameter >= 275:
                baseline = 45.0
            else:
                baseline = 40.0
            
            confidence = "VERY LOW"
            explanation_source = "estimated from size (no historical data)"
    
    # LOAD ACTUAL WOOD DATA FROM EXCEL
    wood_df = load_wood_data()
    
    # Format wood species data for AI
    wood_data_text = ""
    if wood_df is not None and not wood_df.empty:
        wood_data_text = "\nAVAILABLE WOOD SPECIES DATABASE:\n"
        for _, row in wood_df.iterrows():
            species_name = row.get('species', 'Unknown')
            wood_data_text += f"  - {species_name}"
            if 'hardness_category' in row:
                wood_data_text += f": Category={row.get('hardness_category', 'N/A')}"
            if 'base_adjustment_pct' in row:
                wood_data_text += f", Base Adjustment={row.get('base_adjustment_pct', 0):+.1f}%"
            if 'description' in row:
                wood_data_text += f", Description: {row.get('description', '')}"
            wood_data_text += "\n"
    
    # Step 3: AI prediction prompt
    
    prompt = f"""You are a master woodchopping handicapper making precision time predictions for competition.

HANDICAPPING OBJECTIVE

Your prediction must account for wood characteristics and competitor ability to create fair handicaps.
When handicaps are applied, all competitors should finish simultaneously if your predictions are accurate.
This requires deep understanding of how wood properties affect cutting times.

COMPETITOR PROFILE

Name: {competitor_name}
Baseline Time: {baseline:.1f} seconds
Data Source: {explanation_source}
Confidence Level: {confidence}

BASELINE INTERPRETATION:
- This baseline assumes QUALITY 5 wood (average hardness)
- Your task is to adjust this baseline for the ACTUAL quality rating
- Historical data already accounts for competitor's skill level and typical conditions

WOOD SPECIFICATIONS

Species: {species}
Diameter: {diameter:.0f}mm
Quality Rating: {quality}/10
Event Type: {event_code}

WOOD CHARACTERISTICS DATABASE
{wood_data_text}

QUALITY RATING SYSTEM

Quality measures wood condition on a 0-10 scale:

10 = Extremely soft/rotten
   - Wood breaks apart easily
   - Minimal resistance to axe
   - FASTEST possible cutting time
   - Reduces baseline time by approximately 10-15%

9 = Very soft (ideal competition wood)
   - Excellent cutting conditions
   - Clean grain, well-seasoned
   - Reduces baseline time by approximately 7-10%

8 = Soft
   - Good cutting conditions
   - Easy to work with
   - Reduces baseline time by approximately 5-7%

7 = Moderately soft
   - Better than average
   - Noticeable improvement over baseline
   - Reduces baseline time by approximately 3-5%

6 = Slightly soft
   - Marginally better than average
   - Minor improvement
   - Reduces baseline time by approximately 1-3%

5 = AVERAGE HARDNESS (BASELINE REFERENCE POINT)
   - This is what the baseline time assumes
   - NO ADJUSTMENT needed at quality 5
   - Standard competition wood

4 = Slightly hard
   - Marginally tougher than average
   - Minor slowdown
   - Increases baseline time by approximately 1-3%

3 = Moderately hard
   - Noticeably tougher
   - More resistance
   - Increases baseline time by approximately 3-5%

2 = Hard (difficult cutting)
   - Significant resistance
   - Green wood, tough grain
   - Increases baseline time by approximately 5-8%

1 = Very hard
   - Major difficulty
   - Knots, irregular grain
   - Increases baseline time by approximately 8-12%

0 = Extremely hard/barely suitable
   - Maximum difficulty
   - SLOWEST possible cutting time
   - Increases baseline time by approximately 12-15%

CURRENT SITUATION ANALYSIS

Your wood is quality {quality}, which is {abs(quality - 5)} point(s) {"ABOVE" if quality > 5 else "BELOW" if quality < 5 else "AT"} the baseline reference.

{"This wood is SOFTER than baseline - expect FASTER cutting time." if quality > 5 else "This wood is HARDER than baseline - expect SLOWER cutting time." if quality < 5 else "This wood is AVERAGE hardness - baseline time should be accurate."}

CALCULATION METHODOLOGY

Step 1: Start with baseline time: {baseline:.1f}s

Step 2: Apply species base adjustment (if available in database)
- Check species database above for {species}
- Apply the base adjustment percentage if listed

Step 3: Apply quality adjustment
- Calculate deviation from quality 5: {quality} - 5 = {quality - 5}
- Apply adjustment based on quality scale above
- Use the percentage ranges provided (1.5-2.5% per point as guideline)

Step 4: Consider wood physics
- Softer wood (quality >5): Cuts faster, less resistance
- Harder wood (quality <5): Cuts slower, more resistance
- Effect is roughly linear in the middle range (3-7)
- Effect accelerates at extremes (0-2 and 8-10)

Step 5: Validate against typical ranges for {diameter:.0f}mm diameter
- 275mm diameter: 35-50s typical range
- 300mm diameter: 40-55s typical range
- 325mm diameter: 45-60s typical range
- 350mm diameter: 50-70s typical range
- 375mm+ diameter: 60-90s typical range

CRITICAL FACTORS FOR FAIR HANDICAPPING

Front/Back Marker Dynamics:
- Softer wood (quality >5) disproportionately benefits slower competitors (front markers)
  * They gain more time than expected from easier cutting
  * Risk: Front marker finishes before back marker even starts
  
- Harder wood (quality <5) disproportionately penalizes slower competitors
  * They lose more time than expected from difficult cutting
  * Risk: Back marker wins by excessive margin

Your adjustment must account for this to maintain fair handicapping.

WOOD DENSITY AND SIZE INTERACTION

The {diameter:.0f}mm diameter creates a cutting area of approximately {3.14159 * (diameter/2)**2 / 10000:.2f} square cm.
- Larger diameter = exponentially more wood volume to remove
- Quality affects this proportionally: softer wood on large diameter saves significant time
- This diameter/quality interaction is already partially in baseline, but verify your adjustment makes sense

RESPONSE REQUIREMENT

Calculate the most accurate predicted time for {competitor_name} cutting {species} at quality {quality}.

CRITICAL: Respond with ONLY the predicted time as a decimal number.
- Example: 47.3
- NO units (like "seconds" or "s")
- NO explanations
- NO additional text
- JUST THE NUMBER

Predicted time:"""

    response = call_ollama(prompt)
    
    if response is None:
        quality_adjustment = (5 - quality) * 0.02
        predicted_time = baseline * (1 + quality_adjustment)
        explanation = f"Predicted {predicted_time:.1f}s ({explanation_source}, quality adjusted)"
        return predicted_time, confidence, explanation
    
    try:
        import re
        numbers = re.findall(r'\d+\.?\d*', response)
        if numbers:
            predicted_time = float(numbers[0])
            
            if baseline * 0.5 <= predicted_time <= baseline * 1.5:
                explanation = f"Predicted {predicted_time:.1f}s ({explanation_source}, AI quality adjusted)"
                return predicted_time, confidence, explanation
    except:
        pass
    
    explanation = f"Predicted {baseline:.1f}s ({explanation_source})"
    return baseline, confidence, explanation

def calculate_ai_enhanced_handicaps(heat_assignment_df, species, diameter, quality, event_code, results_df, progress_callback=None):
    """
    Calculate handicaps using dual prediction system (Baseline + ML + LLM).

        heat_assignment_df (DataFrame): Competitors in heat
        species (str): Wood species
        diameter (float): Diameter in mm
        quality (int): Quality rating (0-10)
        event_code (str): Event type (SB or UH)
        results_df (DataFrame): Historical results data
        progress_callback (callable): Optional callback function(current, total, competitor_name)

    Returns:
        list: List of dicts with competitor info, ALL predictions (baseline/ML/LLM), and marks
    """
    results = []

    # Ensure quality is an integer
    if quality is None:
        quality = 5
    quality = int(quality)

    total_competitors = len(heat_assignment_df)

    # Run predictions with progress tracking
    for idx, (_, row) in enumerate(heat_assignment_df.iterrows(), 1):
        comp_name = row.get("competitor_name")

        if progress_callback:
            progress_callback(idx, total_competitors, comp_name)

        # Get ALL predictions (baseline, ML, LLM)
        all_preds = get_all_predictions(
            comp_name, species, diameter, quality, event_code, results_df
        )

        # Select best prediction for handicap marks
        predicted_time, method_used, confidence, explanation = select_best_prediction(all_preds)

        if predicted_time is None:
            continue

        results.append({
            'name': comp_name,
            'predicted_time': predicted_time,  # Best prediction for marks
            'method_used': method_used,        # Which method was used
            'confidence': confidence,
            'explanation': explanation,
            'predictions': all_preds           # Store all predictions for display
        })

    # Calculate marks
    if not results:
        return None

    # Sort by predicted time (slowest first)
    results.sort(key=lambda x: x['predicted_time'], reverse=True)

    # Slowest competitor gets mark 3
    slowest_time = results[0]['predicted_time']

    for result in results:
        gap = slowest_time - result['predicted_time']
        mark = 3 + int(gap + 0.999)  # Round up using ceiling logic

        # Apply 180-second maximum rule
        if mark > 183:
            mark = 183

        result['mark'] = mark

    return results



# ============================================================================
# ML PREDICTION SYSTEM (XGBOOST-BASED DUAL PREDICTION)
# ============================================================================

# Global variables to cache trained models (separate for SB and UH)
_cached_ml_model_sb = None
_cached_ml_model_uh = None
_model_training_data_size = 0
_feature_importance_sb = None
_feature_importance_uh = None


def validate_results_data(results_df: pd.DataFrame) -> Tuple[Optional[pd.DataFrame], List[str]]:
    """
    Validate and clean historical results data.

    Args:
        results_df: Raw results DataFrame

    Returns:
        Tuple of (cleaned_df, warnings_list). Returns (None, warnings) if validation fails.
    """
    warnings: List[str] = []

    if results_df is None or results_df.empty:
        return None, ["No data to validate"]

    df = results_df.copy()
    initial_count = len(df)

    # Check for required columns
    required_cols = ['competitor_name', 'event', 'raw_time', 'size_mm', 'species']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        warnings.append(f"Missing columns: {missing_cols}")
        return None, warnings

    # Remove invalid times (use config constants)
    invalid_times = df[
        (df['raw_time'] <= data_req.MIN_VALID_TIME_SECONDS) |
        (df['raw_time'] > data_req.MAX_VALID_TIME_SECONDS)
    ]
    if not invalid_times.empty:
        warnings.append(
            f"Removed {len(invalid_times)} records with impossible times "
            f"(<{data_req.MIN_VALID_TIME_SECONDS}s or >{data_req.MAX_VALID_TIME_SECONDS}s)"
        )
        df = df[
            (df['raw_time'] > data_req.MIN_VALID_TIME_SECONDS) &
            (df['raw_time'] <= data_req.MAX_VALID_TIME_SECONDS)
        ]

    # Remove invalid sizes (use config constants)
    invalid_sizes = df[
        (df['size_mm'] < data_req.MIN_DIAMETER_MM) |
        (df['size_mm'] > data_req.MAX_DIAMETER_MM)
    ]
    if not invalid_sizes.empty:
        warnings.append(
            f"Removed {len(invalid_sizes)} records with invalid diameters "
            f"(<{data_req.MIN_DIAMETER_MM}mm or >{data_req.MAX_DIAMETER_MM}mm)"
        )
        df = df[
            (df['size_mm'] >= data_req.MIN_DIAMETER_MM) &
            (df['size_mm'] <= data_req.MAX_DIAMETER_MM)
        ]

    # Check for missing competitor names
    missing_names = df[df['competitor_name'].isna() | (df['competitor_name'] == '')]
    if not missing_names.empty:
        warnings.append(f"Removed {len(missing_names)} records with missing competitor names")
        df = df[df['competitor_name'].notna() & (df['competitor_name'] != '')]

    # Check for invalid event codes (use config constants)
    invalid_events = df[~df['event'].isin(events.VALID_EVENTS)]
    if not invalid_events.empty:
        warnings.append(
            f"Removed {len(invalid_events)} records with invalid event codes "
            f"(must be {' or '.join(events.VALID_EVENTS)})"
        )
        df = df[df['event'].isin(events.VALID_EVENTS)]

    # Detect statistical outliers using IQR method (per event type)
    outliers_removed = 0
    for event in events.VALID_EVENTS:
        event_data = df[df['event'] == event]['raw_time']
        if len(event_data) > 10:  # Only check if we have enough data
            Q1 = event_data.quantile(0.25)
            Q3 = event_data.quantile(0.75)
            IQR = Q3 - Q1
            # Use config constant for IQR multiplier
            lower_bound = Q1 - data_req.OUTLIER_IQR_MULTIPLIER * IQR
            upper_bound = Q3 + data_req.OUTLIER_IQR_MULTIPLIER * IQR

            outliers = df[(df['event'] == event) &
                         ((df['raw_time'] < lower_bound) | (df['raw_time'] > upper_bound))]
            if not outliers.empty:
                outliers_removed += len(outliers)
                df = df[~((df['event'] == event) &
                         ((df['raw_time'] < lower_bound) | (df['raw_time'] > upper_bound)))]

    if outliers_removed > 0:
        warnings.append(
            f"Removed {outliers_removed} statistical outliers "
            f"(>{data_req.OUTLIER_IQR_MULTIPLIER}x IQR from median)"
        )

    final_count = len(df)
    if final_count < initial_count:
        warnings.append(
            f"Data cleaned: {initial_count} -> {final_count} records "
            f"({initial_count - final_count} removed)"
        )

    if final_count < data_req.MIN_ML_TRAINING_RECORDS_TOTAL:
        warnings.append(
            f"Warning: Only {final_count} valid records remaining "
            f"(need {data_req.MIN_ML_TRAINING_RECORDS_TOTAL}+ for ML training)"
        )

    return df, warnings


def perform_cross_validation(X, y, model_params, cv_folds=5):
    """
    Perform k-fold cross-validation to estimate model accuracy.

    Args:
        X: Feature matrix
        y: Target vector
        model_params: XGBoost parameters dict
        cv_folds: Number of cross-validation folds

    Returns:
        dict: CV results with mean and std of metrics
    """
    if not ML_AVAILABLE:
        return None

    from sklearn.model_selection import cross_val_score

    # Create model with same params as training
    model = xgb.XGBRegressor(**model_params)

    # Perform cross-validation for MAE
    mae_scores = cross_val_score(model, X, y, cv=cv_folds,
                                  scoring='neg_mean_absolute_error',
                                  n_jobs=-1)

    # Perform cross-validation for R²
    r2_scores = cross_val_score(model, X, y, cv=cv_folds,
                                 scoring='r2',
                                 n_jobs=-1)

    return {
        'mae_mean': -mae_scores.mean(),
        'mae_std': mae_scores.std(),
        'r2_mean': r2_scores.mean(),
        'r2_std': r2_scores.std(),
        'mae_scores': -mae_scores,
        'r2_scores': r2_scores
    }


def display_feature_importance(model, event_code, feature_names):
    """
    Display feature importance from trained XGBoost model.

    Args:
        model: Trained XGBoost model
        event_code: 'SB' or 'UH'
        feature_names: List of feature names
    """
    if model is None:
        return

    importance = model.feature_importances_

    # Sort by importance
    importance_pairs = sorted(zip(feature_names, importance),
                             key=lambda x: x[1], reverse=True)

    print(f"\n{'='*70}")
    print(f"  FEATURE IMPORTANCE - {event_code} MODEL")
    print(f"{'='*70}")
    print(f"\n{'Feature':<40} {'Importance':>10}")
    print("-" * 70)

    for name, score in importance_pairs:
        # Create a simple bar chart
        bar_length = int(score * 40)
        bar = '#' * bar_length
        print(f"{name:<40} {score:>9.3f}  {bar}")

    print(f"{'='*70}")

    # Store for later reference
    global _feature_importance_sb, _feature_importance_uh
    if event_code == 'SB':
        _feature_importance_sb = dict(importance_pairs)
    else:
        _feature_importance_uh = dict(importance_pairs)


def engineer_features_for_ml(results_df, wood_df=None):
    """
    Engineer features for ML model from historical results.

    Args:
        results_df: DataFrame with historical results
        wood_df: DataFrame with wood properties (optional, will load if not provided)

    Returns:
        DataFrame with engineered features ready for training/prediction
    """
    if results_df is None or results_df.empty:
        return None

    # Load wood data if not provided
    if wood_df is None:
        wood_df = load_wood_data()

    # Create a copy to avoid modifying original
    df = results_df.copy()

    # Ensure required columns exist
    required_cols = ['competitor_name', 'event', 'raw_time', 'size_mm', 'species']
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        print(f"Warning: Missing required columns for ML: {missing}")
        return None

    # Remove invalid records
    df = df[df['raw_time'] > 0].copy()
    df = df[df['size_mm'] > 0].copy()

    if df.empty:
        return None

    # Feature 1: Event type encoding (SB=0, UH=1)
    df['event_encoded'] = df['event'].apply(lambda x: 0 if str(x).upper() == 'SB' else 1)

    # Feature 2: Competitor average time by event
    competitor_avg = df.groupby(['competitor_name', 'event'])['raw_time'].transform('mean')
    df['competitor_avg_time_by_event'] = competitor_avg

    # Feature 3: Competitor experience (count of past events)
    df['competitor_experience'] = df.groupby('competitor_name').cumcount() + 1

    # Feature 4: Size (already present as size_mm)

    # Feature 5 & 6: Join wood properties (janka_hardness, spec_gravity)
    if not wood_df.empty and 'speciesID' in wood_df.columns:
        # Create species code mapping - use speciesID to match with Results sheet species codes
        wood_properties = wood_df[['speciesID', 'janka_hard', 'spec_gravity']].copy()
        wood_properties = wood_properties.rename(columns={
            'speciesID': 'species',
            'janka_hard': 'wood_janka_hardness',
            'spec_gravity': 'wood_spec_gravity'
        })

        # Join wood properties with results
        df = df.merge(wood_properties, on='species', how='left')

        # Fill missing wood properties with median values (not inplace to avoid warning)
        median_janka = df['wood_janka_hardness'].median()
        median_spec_grav = df['wood_spec_gravity'].median()

        # Use fillna without inplace
        df['wood_janka_hardness'] = df['wood_janka_hardness'].fillna(median_janka if pd.notna(median_janka) else 2000)
        df['wood_spec_gravity'] = df['wood_spec_gravity'].fillna(median_spec_grav if pd.notna(median_spec_grav) else 0.37)
    else:
        # If wood data not available, use default values
        df['wood_janka_hardness'] = 2000  # Default medium hardness
        df['wood_spec_gravity'] = 0.37    # Default medium density

    return df


def train_ml_model(results_df=None, wood_df=None, force_retrain=False, event_code=None):
    """
    Train separate XGBoost models for SB and UH events.

    Args:
        results_df: Historical results DataFrame (will load if not provided)
        wood_df: Wood properties DataFrame (will load if not provided)
        force_retrain: Force retraining even if models are cached
        event_code: Specific event to train ('SB' or 'UH'), or None for both

    Returns:
        dict with 'SB' and 'UH' keys containing trained models, or None if insufficient data
    """
    global _cached_ml_model_sb, _cached_ml_model_uh, _model_training_data_size
    global _feature_importance_sb, _feature_importance_uh

    # Check if ML libraries available
    if not ML_AVAILABLE:
        return None

    # Load data if not provided
    if results_df is None:
        results_df = load_results_df()

    if results_df is None or results_df.empty:
        return None

    # Validate and clean data
    print("\n[DATA VALIDATION]")
    validated_df, warnings = validate_results_data(results_df)

    if warnings:
        print(f"Data validation warnings ({len(warnings)} issues):")
        for i, warning in enumerate(warnings[:5], 1):  # Show first 5
            print(f"  {i}. {warning}")
        if len(warnings) > 5:
            print(f"  ... and {len(warnings) - 5} more")

    if validated_df is None or validated_df.empty:
        print("ERROR: No valid data after validation")
        return None

    print(f"Valid records: {len(validated_df)} / {len(results_df)} ({len(results_df) - len(validated_df)} removed)")

    # Check if we can use cached models
    if not force_retrain and _cached_ml_model_sb is not None and _cached_ml_model_uh is not None:
        if len(validated_df) == _model_training_data_size:
            return {'SB': _cached_ml_model_sb, 'UH': _cached_ml_model_uh}

    # Engineer features
    df_engineered = engineer_features_for_ml(validated_df, wood_df)

    if df_engineered is None or len(df_engineered) < 30:
        print(f"Insufficient data for ML training: {len(df_engineered) if df_engineered is not None else 0} records (need 30+)")
        return None

    # Define features and target
    feature_cols = [
        'competitor_avg_time_by_event',
        'event_encoded',
        'size_mm',
        'wood_janka_hardness',
        'wood_spec_gravity',
        'competitor_experience'
    ]

    # Ensure all feature columns exist
    missing = [col for col in feature_cols if col not in df_engineered.columns]
    if missing:
        print(f"Warning: Missing feature columns: {missing}")
        return None

    # Determine which events to train
    events_to_train = []
    if event_code:
        events_to_train = [event_code.upper()]
    else:
        events_to_train = ['SB', 'UH']

    models = {}

    for event in events_to_train:
        print(f"\n[TRAINING {event} MODEL]")

        # Filter data for this event
        event_df = df_engineered[df_engineered['event'] == event].copy()

        if len(event_df) < 15:  # Reduced minimum for event-specific models
            print(f"Insufficient {event} data: {len(event_df)} records (need 15+)")
            continue

        X = event_df[feature_cols]
        y = event_df['raw_time']

        # Remove any rows with NaN values
        mask = ~(X.isna().any(axis=1) | y.isna())
        X = X[mask]
        y = y[mask]

        if len(X) < 15:
            print(f"Insufficient valid {event} data after cleaning: {len(X)} records")
            continue

        # Model parameters
        model_params = {
            'n_estimators': 100,
            'max_depth': 4,
            'learning_rate': 0.1,
            'random_state': 42,
            'objective': 'reg:squarederror',
            'tree_method': 'hist'
        }

        # Perform cross-validation
        print(f"Cross-validating {event} model (5-fold)...")
        cv_results = perform_cross_validation(X, y, model_params, cv_folds=5)

        if cv_results:
            print(f"  CV MAE: {cv_results['mae_mean']:.2f}s +/- {cv_results['mae_std']:.2f}s")
            print(f"  CV R2:  {cv_results['r2_mean']:.3f} +/- {cv_results['r2_std']:.3f}")

        # Train final model on all data
        model = xgb.XGBRegressor(**model_params)
        model.fit(X, y)

        # Calculate training metrics
        y_pred = model.predict(X)
        mae = mean_absolute_error(y, y_pred)
        r2 = r2_score(y, y_pred)

        print(f"Final {event} model: {len(X)} records (MAE: {mae:.2f}s, R2: {r2:.3f})")

        # Display feature importance
        display_feature_importance(model, event, feature_cols)

        # Cache the model
        if event == 'SB':
            _cached_ml_model_sb = model
            _feature_importance_sb = model.feature_importances_
        else:  # UH
            _cached_ml_model_uh = model
            _feature_importance_uh = model.feature_importances_

        models[event] = model

    # Update training data size
    _model_training_data_size = len(validated_df)

    if not models:
        print("\nERROR: Failed to train any models")
        return None

    # Return both models (use cached if one wasn't trained)
    return {
        'SB': models.get('SB', _cached_ml_model_sb),
        'UH': models.get('UH', _cached_ml_model_uh)
    }


def predict_time_ml(competitor_name, species, diameter, quality, event_code, results_df=None, wood_df=None):
    """
    Predict time using event-specific trained ML model (separate SB/UH models).

    Args:
        competitor_name: Competitor's name
        species: Wood species code
        diameter: Diameter in mm
        quality: Wood quality (0-10) - not used directly but for consistency
        event_code: Event type (SB or UH)
        results_df: Historical results (optional)
        wood_df: Wood properties (optional)

    Returns:
        tuple: (predicted_time, confidence, explanation) or (None, None, None) if error
    """
    global _cached_ml_model_sb, _cached_ml_model_uh

    if not ML_AVAILABLE:
        return None, None, "ML libraries not available"

    # Load data if needed
    if results_df is None:
        results_df = load_results_df()

    if results_df is None or results_df.empty:
        return None, None, "No historical data"

    # Train or get cached models
    models = train_ml_model(results_df, wood_df, force_retrain=False)

    if models is None:
        return None, None, "ML model training failed"

    # Select the appropriate model for this event
    event_upper = event_code.upper()
    if event_upper not in models or models[event_upper] is None:
        return None, None, f"No {event_upper} model available"

    model = models[event_upper]

    # Load wood data for properties
    if wood_df is None:
        wood_df = load_wood_data()

    # Get wood properties
    wood_janka = 500  # Default
    wood_spec_grav = 0.5  # Default

    if not wood_df.empty:
        wood_row = wood_df[wood_df['species'] == species]
        if not wood_row.empty:
            wood_janka = wood_row.iloc[0].get('janka_hard', 500)
            wood_spec_grav = wood_row.iloc[0].get('spec_gravity', 0.5)

    # Calculate competitor average time for this event
    comp_data = results_df[
        (results_df['competitor_name'] == competitor_name) &
        (results_df['event'] == event_code)
    ]

    if not comp_data.empty:
        competitor_avg = comp_data['raw_time'].mean()
        experience = len(comp_data)
        confidence = "HIGH" if len(comp_data) >= 5 else "MEDIUM"
    else:
        # Fallback: use all competitor data regardless of event
        comp_all_data = results_df[results_df['competitor_name'] == competitor_name]
        if not comp_all_data.empty:
            competitor_avg = comp_all_data['raw_time'].mean()
            experience = len(comp_all_data)
            confidence = "MEDIUM"
        else:
            # New competitor: use event average
            event_data = results_df[results_df['event'] == event_code]
            if not event_data.empty:
                competitor_avg = event_data['raw_time'].mean()
                experience = 1
                confidence = "LOW"
            else:
                return None, None, "No reference data for prediction"

    # Prepare features
    event_encoded = 0 if event_code.upper() == 'SB' else 1

    features = pd.DataFrame({
        'competitor_avg_time_by_event': [competitor_avg],
        'event_encoded': [event_encoded],
        'size_mm': [diameter],
        'wood_janka_hardness': [wood_janka],
        'wood_spec_gravity': [wood_spec_grav],
        'competitor_experience': [experience]
    })

    # Make prediction using event-specific model
    try:
        predicted_time = model.predict(features)[0]

        # Sanity check: ensure prediction is reasonable
        if predicted_time < 5 or predicted_time > 300:
            return None, None, f"ML prediction out of range ({predicted_time:.1f}s)"

        explanation = f"{event_upper} ML model ({_model_training_data_size} training records)"

        return predicted_time, confidence, explanation

    except Exception as e:
        return None, None, f"ML prediction error: {str(e)}"


def get_all_predictions(competitor_name, species, diameter, quality, event_code, results_df=None):
    """
    Get predictions from all three methods: Baseline, ML, and LLM.

    Returns:
        dict with keys 'baseline', 'ml', 'llm', each containing:
            {'time': float, 'confidence': str, 'explanation': str, 'error': str or None}
    """
    predictions = {
        'baseline': {'time': None, 'confidence': None, 'explanation': None, 'error': None},
        'ml': {'time': None, 'confidence': None, 'explanation': None, 'error': None},
        'llm': {'time': None, 'confidence': None, 'explanation': None, 'error': None}
    }

    # Load results once
    if results_df is None:
        results_df = load_results_df()

    # 1. Get baseline prediction (existing logic from predict_competitor_time_with_ai)
    # This is the statistical baseline without quality adjustment
    historical_times, data_source = get_competitor_historical_times_flexible(
        competitor_name, species, event_code, results_df
    )

    if len(historical_times) >= 3:
        weights = [1.5, 1.3, 1.1, 1.0, 0.9, 0.8]
        weighted_times = [t * w for t, w in zip(historical_times[:6], weights[:len(historical_times)])]
        baseline = sum(weighted_times) / sum(weights[:len(historical_times)])
        confidence = "HIGH"
        explanation = f"Statistical baseline ({data_source})"
    elif len(historical_times) > 0:
        baseline = statistics.mean(historical_times)
        confidence = "MEDIUM"
        explanation = f"Limited history ({data_source})"
    else:
        baseline, baseline_source = get_event_baseline_flexible(species, diameter, event_code, results_df)
        if baseline:
            confidence = "LOW"
            explanation = f"Event baseline ({baseline_source})"
        else:
            # Ultimate fallback based on diameter
            if diameter >= 350:
                baseline = 60.0
            elif diameter >= 300:
                baseline = 45.0
            elif diameter >= 250:
                baseline = 35.0
            else:
                baseline = 30.0
            confidence = "LOW"
            explanation = "Default estimate (no history)"

    predictions['baseline']['time'] = baseline
    predictions['baseline']['confidence'] = confidence
    predictions['baseline']['explanation'] = explanation

    # 2. Get ML prediction
    ml_time, ml_conf, ml_expl = predict_time_ml(
        competitor_name, species, diameter, quality, event_code, results_df
    )

    if ml_time is not None:
        predictions['ml']['time'] = ml_time
        predictions['ml']['confidence'] = ml_conf
        predictions['ml']['explanation'] = ml_expl
    else:
        predictions['ml']['error'] = ml_expl if ml_expl else "ML prediction unavailable"

    # 3. Get LLM prediction (call existing function)
    llm_time, llm_conf, llm_expl = predict_competitor_time_with_ai(
        competitor_name, species, diameter, quality, event_code, results_df
    )

    if llm_time is not None:
        predictions['llm']['time'] = llm_time
        predictions['llm']['confidence'] = llm_conf
        predictions['llm']['explanation'] = llm_expl
    else:
        predictions['llm']['error'] = "LLM prediction failed"

    return predictions


def select_best_prediction(all_predictions):
    """
    Select the best prediction from available methods.
    Priority: ML > LLM > Baseline

    Args:
        all_predictions: dict from get_all_predictions()

    Returns:
        tuple: (predicted_time, method_name, confidence, explanation)
    """
    # Priority 1: ML (if available and valid)
    if all_predictions['ml']['time'] is not None:
        return (
            all_predictions['ml']['time'],
            'ML',
            all_predictions['ml']['confidence'],
            all_predictions['ml']['explanation']
        )

    # Priority 2: LLM (if available and valid)
    if all_predictions['llm']['time'] is not None:
        return (
            all_predictions['llm']['time'],
            'LLM',
            all_predictions['llm']['confidence'],
            all_predictions['llm']['explanation']
        )

    # Priority 3: Baseline (always available)
    return (
        all_predictions['baseline']['time'],
        'Baseline',
        all_predictions['baseline']['confidence'],
        all_predictions['baseline']['explanation']
    )


def generate_prediction_analysis_llm(all_competitors_predictions, wood_selection):
    """
    Use LLM to analyze differences between ML and LLM predictions across all competitors.

    Args:
        all_competitors_predictions: list of dicts with competitor predictions
        wood_selection: dict with wood characteristics

    Returns:
        str: Natural language analysis
    """
    if not call_ollama("test", model="qwen2.5:7b"):
        return "LLM analysis unavailable (Ollama not running)"

    # Build concise summary for LLM
    summary_lines = []
    for comp_pred in all_competitors_predictions[:10]:  # Limit to 10 for prompt size
        name = comp_pred['name'][:20]  # Truncate long names
        baseline = comp_pred['predictions']['baseline']['time']
        ml = comp_pred['predictions']['ml']['time']
        llm = comp_pred['predictions']['llm']['time']

        if ml and llm:
            ml_llm_diff = ml - llm
            summary_lines.append(f"{name}: Baseline={baseline:.1f}s, ML={ml:.1f}s, LLM={llm:.1f}s (diff={ml_llm_diff:+.1f}s)")
        elif ml:
            summary_lines.append(f"{name}: Baseline={baseline:.1f}s, ML={ml:.1f}s, LLM=N/A")
        elif llm:
            summary_lines.append(f"{name}: Baseline={baseline:.1f}s, ML=N/A, LLM={llm:.1f}s")

    summary_text = "\n".join(summary_lines)

    prompt = f"""You are analyzing handicap predictions for a woodchopping competition.

WOOD CHARACTERISTICS:
- Species: {wood_selection.get('species', 'Unknown')}
- Diameter: {wood_selection.get('size_mm', 0)}mm
- Quality: {wood_selection.get('quality', 5)}/10
- Event: {wood_selection.get('event', 'Unknown')}

PREDICTION COMPARISON (Baseline vs ML vs LLM):
{summary_text}

Analyze these predictions and provide a brief 3-4 sentence analysis covering:
1. Overall agreement or divergence between methods
2. Possible reasons for any significant differences
3. Which method appears most reliable for this scenario
4. Any recommendations for the judge

Keep your response concise and practical."""

    response = call_ollama(prompt, model="qwen2.5:7b")

    if response:
        return response
    else:
        return "Unable to generate LLM analysis at this time."


def display_dual_predictions(handicap_results, wood_selection):
    """
    Display handicap marks with all three prediction methods side-by-side.

    Args:
        handicap_results: List of dicts from calculate_ai_enhanced_handicaps()
        wood_selection: Dict with wood characteristics

    Shows:
        - Competitor name
        - Handicap mark
        - Baseline prediction
        - ML prediction
        - LLM prediction
        - Method used for marks
        - Summary of methods
        - Optional AI analysis of differences
    """
    if not handicap_results:
        print("No handicap results to display")
        return

    # Sort by mark (ascending)
    sorted_results = sorted(handicap_results, key=lambda x: x['mark'])

    # Build header
    print("\n" + "=" * 110)
    wood_info = f"{wood_selection.get('species', 'Unknown')}, {wood_selection.get('size_mm', 0)}mm, Quality: {wood_selection.get('quality', 5)}"
    print(f"HANDICAP MARKS - {wood_info}")
    print("=" * 110)

    # Column headers
    print(f"\n{'Competitor Name':<35} {'Mark':>4}  {'Baseline':>9}  {'ML Model':>9}  {'LLM Model':>9}  {'Used':<8}")
    print("-" * 110)

    # Count methods available
    ml_available_count = 0
    llm_available_count = 0
    method_counts = {'Baseline': 0, 'ML': 0, 'LLM': 0}

    # Display each competitor
    for comp in sorted_results:
        name = comp['name'][:35]
        mark = comp['mark']

        # Get predictions
        baseline_time = comp['predictions']['baseline']['time']
        ml_time = comp['predictions']['ml']['time']
        llm_time = comp['predictions']['llm']['time']

        # Format predictions (show "N/A" if None)
        baseline_str = f"{baseline_time:.1f}s" if baseline_time else "N/A"
        ml_str = f"{ml_time:.1f}s" if ml_time else "N/A"
        llm_str = f"{llm_time:.1f}s" if llm_time else "N/A"

        # Track which method was used
        method_used = comp.get('method_used', 'Unknown')
        method_counts[method_used] = method_counts.get(method_used, 0) + 1

        # Count availability
        if ml_time is not None:
            ml_available_count += 1
        if llm_time is not None:
            llm_available_count += 1

        print(f"{name:<35} {mark:4d}  {baseline_str:>9}  {ml_str:>9}  {llm_str:>9}  {method_used:<8}")

    print("=" * 110)

    # Display prediction methods summary
    print("\nPrediction Methods Summary:")
    print(f"  • Baseline: Statistical calculation (always available)")

    if ml_available_count > 0:
        ml_status = "HIGH" if _model_training_data_size >= 80 else "MEDIUM" if _model_training_data_size >= 50 else "LOW"
        print(f"  • ML Model: XGBoost trained on {_model_training_data_size} records [CONFIDENCE: {ml_status}] - Available for {ml_available_count}/{len(sorted_results)} competitors")
    else:
        print(f"  • ML Model: Not available (insufficient training data)")

    if llm_available_count > 0:
        print(f"  • LLM Model: Ollama qwen2.5:7b AI prediction - Available for {llm_available_count}/{len(sorted_results)} competitors")
    else:
        print(f"  • LLM Model: Not available (Ollama not running or prediction failed)")

    # Show which method was primarily used
    primary_method = max(method_counts, key=method_counts.get)
    print(f"\nMarks calculated using: {primary_method} predictions")
    print(f"(Priority: ML > LLM > Baseline - most accurate method available)")

    # Offer AI analysis
    print("\n" + "=" * 110)
    analyze = input("\nPress Enter to see AI analysis of prediction differences (or 'n' to skip): ").strip().lower()

    if analyze != 'n':
        print("\n" + "=" * 110)
        print("AI ANALYSIS OF PREDICTIONS")
        print("=" * 110)
        print("\nAnalyzing prediction differences...")

        analysis = generate_prediction_analysis_llm(handicap_results, wood_selection)

        # Word wrap the analysis for better readability
        wrapped_lines = textwrap.wrap(analysis, width=106)
        for line in wrapped_lines:
            print(line)

        print("\n" + "=" * 110)


# MONTE CARLO SIMULATION FUNCTIONS (USED BY MENU 4)

""" Simulate a single race with realistic ABSOLUTE performance variation.

    competitors_with_marks (list): List of dicts with name, predicted_time, mark

Returns:
    list: Finish results sorted by finish time """


## Simulate a singular race with absolute variance: ±3 seconds standard deviation for everyone
'''3.0 Defines the standard deviation for time variation

~68% of the time, actual time will be within ±3 seconds of predicted
~95% of the time, actual time will be within ±6 seconds of predicted
This models real-world unpredictability'''

def simulate_single_race(competitors_with_marks: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Simulate a single race with absolute performance variation.

    Args:
        competitors_with_marks: List of competitor dicts with marks and predicted times

    Returns:
        List of finish results sorted by finish time
    """
    finish_results = []

    for comp in competitors_with_marks:
        # Apply absolute variance (±3 seconds for all competitors - critical for fairness)
        actual_time = np.random.normal(
            comp['predicted_time'],
            rules.PERFORMANCE_VARIANCE_SECONDS
        )

        # Prevent unreasonably fast times
        actual_time = max(actual_time, comp['predicted_time'] * 0.5)

        # Calculate finish time accounting for handicap
        start_delay = comp['mark'] - rules.MIN_MARK_SECONDS  # Front marker starts immediately
        finish_time = start_delay + actual_time
        
        finish_results.append({
            'name': comp['name'],
            'mark': comp['mark'],
            'actual_time': actual_time,
            'finish_time': finish_time,
            'predicted_time': comp['predicted_time']
        })
    
    # Sort by finish time
    finish_results.sort(key=lambda x: x['finish_time'])
    
    return finish_results


##Monte Carlo Simulation Function

'''Run Monte Carlo simulation to assess handicap fairness.'''

def run_monte_carlo_simulation(
    competitors_with_marks: List[Dict[str, Any]],
    num_simulations: int = None
) -> Dict[str, Any]:
    """
    Run Monte Carlo simulation to assess handicap fairness.

    Args:
        competitors_with_marks: List of competitor dicts with marks and predicted times
        num_simulations: Number of race simulations to run (defaults to config value)

    Returns:
        Dictionary containing comprehensive simulation statistics
    """
    if num_simulations is None:
        num_simulations = sim_config.NUM_SIMULATIONS

    print("\n" + "=" * display.SEPARATOR_LENGTH)
    print(f"RUNNING MONTE CARLO SIMULATION ({num_simulations:,} races)")
    print("=" * display.SEPARATOR_LENGTH)
    print(f"Simulating races with ±{rules.PERFORMANCE_VARIANCE_SECONDS} second absolute performance variation...")
    
    # Track statistics
    finish_spreads = []
    winner_counts = {comp['name']: 0 for comp in competitors_with_marks}
    podium_counts = {comp['name']: 0 for comp in competitors_with_marks}  # Top 3
    finish_position_sums = {comp['name']: 0 for comp in competitors_with_marks}
    
    # Track front marker (slowest predicted, starts first)
    front_marker_name = max(competitors_with_marks, key=lambda x: x['predicted_time'])['name']
    back_marker_name = min(competitors_with_marks, key=lambda x: x['predicted_time'])['name']
    
    # Run simulations
    for i in range(num_simulations):
        if (i + 1) % 50000 == 0:
            print(f"  Completed {i + 1}/{num_simulations} simulations...")
        
        race_results = simulate_single_race(competitors_with_marks)
        
        # Calculate finish spread
        spread = race_results[-1]['finish_time'] - race_results[0]['finish_time']
        finish_spreads.append(spread)
        
        # Track winner
        winner_counts[race_results[0]['name']] += 1
        
        # Track podium (top 3)
        for j in range(min(3, len(race_results))):
            podium_counts[race_results[j]['name']] += 1
        
        # Track average finish positions
        for pos, result in enumerate(race_results, 1):
            finish_position_sums[result['name']] += pos
    
    # Calculate statistics
    avg_finish_positions = {name: pos_sum / num_simulations 
                           for name, pos_sum in finish_position_sums.items()}
    
    analysis = {
        'num_simulations': num_simulations,
        'finish_spreads': finish_spreads,
        'avg_spread': np.mean(finish_spreads),
        'median_spread': np.median(finish_spreads),
        'min_spread': np.min(finish_spreads),
        'max_spread': np.max(finish_spreads),
        'tight_finish_prob': sum(1 for s in finish_spreads if s < 10) / num_simulations,
        'very_tight_finish_prob': sum(1 for s in finish_spreads if s < 5) / num_simulations,
        'winner_counts': winner_counts,
        'winner_percentages': {name: (count / num_simulations * 100) 
                              for name, count in winner_counts.items()},
        'podium_counts': podium_counts,
        'podium_percentages': {name: (count / num_simulations * 100) 
                              for name, count in podium_counts.items()},
        'avg_finish_positions': avg_finish_positions,
        'front_marker_name': front_marker_name,
        'back_marker_name': back_marker_name,
        'front_marker_wins': winner_counts[front_marker_name],
        'back_marker_wins': winner_counts[back_marker_name],
        'competitors': competitors_with_marks
    }
    
    return analysis

#Generate summary of the analysis of the Monte Carlo simulation

def generate_simulation_summary(analysis):
    
    summary = []
    summary.append("\n" + "="*70)
    summary.append("MONTE CARLO SIMULATION RESULTS")
    summary.append("="*70)
    summary.append(f"Simulated {analysis['num_simulations']} races with ±3s absolute performance variation")
    summary.append("")
    
    summary.append("FINISH TIME SPREADS:")
    summary.append(f"  Average spread: {analysis['avg_spread']:.1f} seconds")
    summary.append(f"  Median spread:  {analysis['median_spread']:.1f} seconds")
    summary.append(f"  Range: {analysis['min_spread']:.1f}s - {analysis['max_spread']:.1f}s")
    summary.append(f"  Tight finish (<10s): {analysis['tight_finish_prob']*100:.1f}% of races")
    summary.append(f"  Very tight (<5s):    {analysis['very_tight_finish_prob']*100:.1f}% of races")
    summary.append("")
    
    summary.append("WIN PROBABILITIES:")
    sorted_winners = sorted(analysis['winner_percentages'].items(), 
                           key=lambda x: x[1], reverse=True)
    for name, pct in sorted_winners:
        summary.append(f"  {name:25s} {pct:5.1f}% ({analysis['winner_counts'][name]:4d} wins)")
    summary.append("")
    
    summary.append("AVERAGE FINISH POSITIONS:")
    sorted_positions = sorted(analysis['avg_finish_positions'].items(), 
                             key=lambda x: x[1])
    for name, avg_pos in sorted_positions:
        summary.append(f"  {name:25s} Avg position: {avg_pos:.2f}")
    summary.append("")
    
    summary.append("FRONT/BACK MARKER ANALYSIS:")
    summary.append(f"  Front marker (slowest): {analysis['front_marker_name']}")
    summary.append(f"    Win rate: {analysis['front_marker_wins']/analysis['num_simulations']*100:.1f}%")
    summary.append(f"  Back marker (fastest): {analysis['back_marker_name']}")
    summary.append(f"    Win rate: {analysis['back_marker_wins']/analysis['num_simulations']*100:.1f}%")
    summary.append("="*70)
    
    return "\n".join(summary)

#gENERATE cHART SHOWING WIN-RATES
def visualize_simulation_results(analysis):

    print("\n" + "="*70)
    print("WIN RATE VISUALIZATION")
    print("="*70)
    
    max_pct = max(analysis['winner_percentages'].values())
    
    sorted_winners = sorted(analysis['winner_percentages'].items(), 
                           key=lambda x: x[1], reverse=True)
    
    for name, pct in sorted_winners:
        bar_length = int((pct / max_pct) * 40)  # Scale to 40 chars max
        bar = "█" * bar_length
        print(f"{name:25s} {pct:5.1f}% {bar}")
    
    print("="*70)


#Assess the fairness of handicaps
''' Use LLM to provide expert assessment of handicap fairness. '''

def get_ai_assessment_of_handicaps(analysis):
    
    # Calculate fairness metrics
    max_win_rate = max(analysis['winner_percentages'].values())
    min_win_rate = min(analysis['winner_percentages'].values())
    win_rate_spread = max_win_rate - min_win_rate
    ideal_win_rate = 100.0 / len(analysis['competitors'])
    
    # Calculate per-competitor deviations
    win_rate_deviations = {}
    for name, pct in analysis['winner_percentages'].items():
        deviation = pct - ideal_win_rate
        win_rate_deviations[name] = deviation
    
    # Identify extremes
    most_favored = max(win_rate_deviations, key=win_rate_deviations.get)
    most_disadvantaged = min(win_rate_deviations, key=win_rate_deviations.get)
    
    # Format data for prompt
    winner_data = "\n".join([f"  - {name}: {pct:.2f}% win rate (deviation: {win_rate_deviations[name]:+.2f}%)" 
                            for name, pct in sorted(analysis['winner_percentages'].items(), 
                                                   key=lambda x: x[1], reverse=True)])
    
    competitor_details = "\n".join([f"  - {comp['name']}: {comp['predicted_time']:.1f}s predicted → Mark {comp['mark']}"
                                    for comp in sorted(analysis['competitors'], 
                                                      key=lambda x: x['predicted_time'], reverse=True)])
    
    win_rate_std_dev = np.std(list(analysis['winner_percentages'].values()))
    coefficient_of_variation = (win_rate_std_dev / ideal_win_rate) * 100 if ideal_win_rate > 0 else 0
    
    prompt = f"""You are a master woodchopping handicapper and statistician analyzing the fairness of predicted handicap marks through Monte Carlo simulation.

HANDICAPPING PRINCIPLES

PRIMARY GOAL: Create handicaps where ALL competitors have EQUAL probability of winning.
- In a fair handicap system, skill level should NOT predict victory
- A novice with Mark 3 should win as often as an expert with Mark 25
- The slowest competitor should have the same chance as the fastest

HANDICAPPING MECHANISM:
1. Predict each competitor's raw cutting time
2. Slowest predicted time receives Mark 3 (starts first)
3. Faster predicted times receive higher marks (delayed starts)
4. If predictions are perfect, everyone finishes simultaneously
5. Natural variation (±3s) creates competitive spread

QUALITY FACTORS IN PREDICTIONS:
- Wood species (hardness variations)
- Block diameter (volume to cut)
- Wood quality rating (0-10 scale, affects cutting speed)
- Historical competitor performance

SIMULATION METHODOLOGY

WHAT WE TESTED:
- Simulated {analysis['num_simulations']:,} races with {len(analysis['competitors'])} competitors
- Applied ±3 second ABSOLUTE performance variation (realistic race conditions)
- Variation represents: technique consistency, wood grain, fatigue, environmental conditions

WHY ABSOLUTE VARIANCE (±3s for everyone):
- Real factors affect all skill levels equally in absolute seconds
- Wood grain knot costs 2s for novice AND expert (not proportional to skill)
- Technique wobble affects everyone by similar absolute time
- This is a CRITICAL breakthrough in fair handicapping

STATISTICAL SIGNIFICANCE:
- With {analysis['num_simulations']:,} simulations, margin of error is extremely small
- Patterns in results are REAL, not random noise
- Even 1-2% win rate differences are statistically meaningful

SIMULATION RESULTS

COMPETITOR PREDICTIONS AND MARKS:
{competitor_details}

IDEAL WIN RATE: {ideal_win_rate:.2f}% per competitor
(Perfect handicapping means all competitors win exactly {ideal_win_rate:.2f}% of races)

ACTUAL WIN RATES:
{winner_data}

STATISTICAL MEASURES:
- Win Rate Spread: {win_rate_spread:.2f}% (maximum minus minimum)
- Standard Deviation: {win_rate_std_dev:.2f}%
- Coefficient of Variation: {coefficient_of_variation:.1f}%

FINISH TIME ANALYSIS:
- Average finish spread: {analysis['avg_spread']:.1f} seconds
- Median finish spread: {analysis['median_spread']:.1f} seconds
- Tight finishes (<10s): {analysis['tight_finish_prob']*100:.1f}% of races
- Very tight finishes (<5s): {analysis['very_tight_finish_prob']*100:.1f}% of races

FRONT AND BACK MARKER PERFORMANCE:
- Front Marker (slowest): {analysis['front_marker_name']} - {analysis['front_marker_wins']/analysis['num_simulations']*100:.1f}% wins
- Back Marker (fastest): {analysis['back_marker_name']} - {analysis['back_marker_wins']/analysis['num_simulations']*100:.1f}% wins

PATTERN IDENTIFICATION:
- Most Favored: {most_favored} ({analysis['winner_percentages'][most_favored]:.2f}%, +{win_rate_deviations[most_favored]:.2f}%)
- Most Disadvantaged: {most_disadvantaged} ({analysis['winner_percentages'][most_disadvantaged]:.2f}%, {win_rate_deviations[most_disadvantaged]:.2f}%)

FAIRNESS CRITERIA

RATING SCALE (based on win rate spread):

EXCELLENT (Spread ≤ 3%):
- All win rates within ±1.5% of ideal ({ideal_win_rate-1.5:.1f}% to {ideal_win_rate+1.5:.1f}%)
- Handicaps are nearly perfect
- Predictions are highly accurate
- No adjustments needed

VERY GOOD (Spread ≤ 6%):
- All win rates within ±3% of ideal ({ideal_win_rate-3:.1f}% to {ideal_win_rate+3:.1f}%)
- Handicaps are working well
- Minor prediction inaccuracies
- Only minor adjustments if desired

GOOD (Spread ≤ 10%):
- All win rates within ±5% of ideal ({ideal_win_rate-5:.1f}% to {ideal_win_rate+5:.1f}%)
- Acceptable fairness for competition
- Some prediction bias exists
- Consider adjustments for championship events

FAIR (Spread ≤ 16%):
- Win rates within ±8% of ideal
- Noticeable imbalance
- Predictions need refinement
- Adjustments recommended

POOR (Spread > 16%):
- Significant imbalance detected
- Predictions are systematically biased
- Handicaps require major adjustment
- Not suitable for fair competition

UNACCEPTABLE (Any competitor >2x or <0.5x ideal):
- Extreme bias detected
- One competitor has double (or half) expected win rate
- Fundamental prediction error
- Complete recalibration required

DIAGNOSTIC PATTERNS

COMMON ISSUES TO IDENTIFY:

1. FRONT MARKER ADVANTAGE (soft wood bias):
   Pattern: Front marker wins >ideal, back marker wins <ideal
   Cause: Predictions underestimate benefit of soft wood to slower competitors
   Fix: Increase quality adjustment for front markers on soft wood

2. BACK MARKER ADVANTAGE (hard wood bias):
   Pattern: Back marker wins >ideal, front marker wins <ideal
   Cause: Predictions underestimate difficulty of hard wood for slower competitors
   Fix: Increase time penalties for front markers on hard wood

3. MIDDLE COMPRESSION:
   Pattern: Extreme competitors (fastest/slowest) win less than middle competitors
   Cause: Predictions too conservative at extremes
   Fix: Increase handicap spread (widen gaps between marks)

4. EXPERIENCE BIAS:
   Pattern: Competitors with more historical data win more often
   Cause: Better predictions for experienced competitors
   Fix: Adjust confidence weighting or baseline calculations

5. SPECIES MISCALIBRATION:
   Pattern: Systematic bias across all competitors
   Cause: Species hardness factor incorrect
   Fix: Adjust species baseline percentage

YOUR ANALYSIS TASK

Provide a comprehensive assessment in the following structure:

1. FAIRNESS RATING: State one of: Excellent / Very Good / Good / Fair / Poor / Unacceptable

2. STATISTICAL ANALYSIS (2-3 sentences):
   - Interpret the win rate spread of {win_rate_spread:.2f}%
   - Comment on finish time spreads (average {analysis['avg_spread']:.1f}s)
   - Assess if variation is appropriate for exciting competition

3. PATTERN DIAGNOSIS (2-3 sentences):
   - Identify which diagnostic pattern (if any) is present
   - Explain WHY this pattern occurred based on competitor times
   - Reference specific competitors showing the bias

4. PREDICTION ACCURACY (1-2 sentences):
   - Are the predictions systematically biased or just slightly off?
   - Is the issue with one competitor or system-wide?

5. RECOMMENDATIONS (2-3 specific actions):
   If EXCELLENT or VERY GOOD: Affirm handicaps are ready for use
   If GOOD: Suggest optional refinements
   If FAIR, POOR, or UNACCEPTABLE: Provide specific adjustment recommendations
   
   Format recommendations as bullet points:
   • First specific action (include numbers when possible)
   • Second specific action
   • Final recommendation

RESPONSE REQUIREMENTS:
- Keep total response to 8-12 sentences maximum
- Be specific and actionable
- Use technical terms confidently
- Cite actual numbers from the data above
- Base analysis on ACTUAL DATA, not generic observations
- Reference specific competitors, percentages, and patterns you observe

Your Expert Assessment:"""

    response = call_ollama(prompt)
    
    if response:
        return response
    else:
        # Enhanced fallback assessment
        if win_rate_spread < 3:
            rating = "EXCELLENT"
            assessment = "Handicaps are nearly perfect. Predictions are highly accurate with minimal bias."
        elif win_rate_spread < 6:
            rating = "VERY GOOD"
            assessment = "Handicaps are working very well. Minor prediction variations are within acceptable range."
        elif win_rate_spread < 10:
            rating = "GOOD"
            assessment = "Handicaps are acceptable for competition. Some prediction refinement would improve fairness."
        elif win_rate_spread < 16:
            rating = "FAIR"
            assessment = "Noticeable imbalance detected. Predictions show systematic bias requiring adjustment."
        else:
            rating = "POOR"
            assessment = "Significant imbalance requiring major prediction recalibration."
        
        front_wins = analysis['front_marker_wins']/analysis['num_simulations']*100
        back_wins = analysis['back_marker_wins']/analysis['num_simulations']*100
        
        if front_wins > ideal_win_rate + 3:
            pattern = "Front marker advantage detected (soft wood bias likely)."
        elif back_wins > ideal_win_rate + 3:
            pattern = "Back marker advantage detected (hard wood bias likely)."
        else:
            pattern = "No clear front/back marker bias pattern."
        
        return f"""FAIRNESS RATING: {rating}

STATISTICAL ANALYSIS: With {len(analysis['competitors'])} competitors, ideal win rate is {ideal_win_rate:.1f}% each. Actual spread is {win_rate_spread:.2f}% (from {min_win_rate:.1f}% to {max_win_rate:.1f}%). {assessment} Average finish spread of {analysis['avg_spread']:.1f}s creates exciting competition.

PATTERN DIAGNOSIS: {pattern} {most_favored} is most favored at {analysis['winner_percentages'][most_favored]:.1f}% wins (+{win_rate_deviations[most_favored]:.1f}% above ideal), while {most_disadvantaged} is disadvantaged at {analysis['winner_percentages'][most_disadvantaged]:.1f}% wins ({win_rate_deviations[most_disadvantaged]:.1f}% below ideal).

RECOMMENDATIONS:
- {"Handicaps are ready for competition use - no adjustments needed." if win_rate_spread < 6 else f"Review predictions for {most_favored} and {most_disadvantaged} - time estimates may need adjustment."}
- {"Continue collecting historical data to improve future predictions." if win_rate_spread < 10 else "Consider adjusting quality/species factors in prediction model."}
- {"Monitor real competition results to validate simulation predictions." if win_rate_spread < 16 else "Recalibrate baseline calculations before using these handicaps in competition."}"""

#Main function to simulate the handicaps and assess the fairness

def simulate_and_assess_handicaps(competitors_with_marks, num_simulations=None):

    if not competitors_with_marks or len(competitors_with_marks) < 2:
        print("Need at least 2 competitors to run simulation.")
        return

    # Use config default if not specified
    if num_simulations is None:
        num_simulations = sim_config.NUM_SIMULATIONS

    # Run simulation
    analysis = run_monte_carlo_simulation(competitors_with_marks, num_simulations)
    
    # Display results
    summary = generate_simulation_summary(analysis)
    print(summary)
    
    # Visualize
    visualize_simulation_results(analysis)
    
    # Get AI assessment
    print("\n" + "="*70)
    print("AI HANDICAPPING ASSESSMENT")
    print("="*70)
    print("\nAnalyzing fairness of handicaps...")
    
    ai_assessment = get_ai_assessment_of_handicaps(analysis)

    # Preserve structure while wrapping long lines
    # Split into paragraphs first, then wrap each paragraph independently
    print("")  # Add spacing
    paragraphs = ai_assessment.split('\n\n')
    for paragraph in paragraphs:
        if paragraph.strip():
            # Check if it's a section header (contains colon at start)
            lines = paragraph.split('\n')
            for line in lines:
                if line.strip():
                    # Wrap long lines to 100 characters for better readability
                    wrapped_lines = textwrap.wrap(line, width=100, subsequent_indent='  ')
                    for wrapped_line in wrapped_lines:
                        print(wrapped_line)
            print()  # Blank line between sections

    print("="*70)



# MENU OPTION 4: VIEW HANDICAP MARKS (AI-ENHANCED)
''' Official will be presented with the calculated handicap marks for each selected competitor in the heat
    1. View Handicap Marks
    2. Return to Main Menu

 3 Main Functions:
        -One to view the handicap menu
        -One to validate the data
        -One to actually tabulate the data

        1) Official finishes running a heat
        2) Calls this function with competitor list and wood specs
        3) Function validates event is selected
        4) Prompts for Heat ID (optional)
        4) Loops through each competitor asking for their time
        5) Skips anyone with blank time entry
        6) Converts valid times to proper format and collects in list
        7) Opens Excel file (or creates new one)
        8) Finds/creates Results sheet
        9) Write all result rows at once
        10) Save and close file

'''



#View Handicap Marks Menu
def view_handicaps_menu(heat_assignment_df, wood_selection):
    ''' Official will be presented with the calculated handicap marks for each selected competitor in the heat
        1. View handicap marks for current heat
        2. Return to Main Menu
    '''

    #safety check to make sure that eiyther standing block or Underhand is selected. If not, will default to none so program doesn't crash
    if "event" not in wood_selection:
        wood_selection["event"] = None

    #menu loop
    while True:
        print("\n--- View Handicap Marks ---")
        print("1) View handicap marks for current heat")
        print("2) Back to Main Menu")
        s = input("Choose an option: ").strip()

        if s == "1":
            if not validate_heat_data(heat_assignment_df, wood_selection):
                continue
            view_handicaps(heat_assignment_df, wood_selection)
            input("\n(Press Enter to return to the View Handicap Marks menu) ")

        elif s == "2" or s == "":
            break

        else:
            print("Invalid selection. Try again.")

#Helper to validate heat data is complete

    '''1)CHecks if all data exists
    2) jumps back to menu if data is incomplete

    Checks specifically for: competitors in heat assignment, species, size, and event"
    '''

def validate_heat_data(heat_assignment_df, wood_selection):
    if heat_assignment_df is None or heat_assignment_df.empty:
        print("\nNo competitors in heat assignment. Use Competitor Menu -> Select Competitors for Heat.")
        return False
    
    if not wood_selection.get("species") or not wood_selection.get("size_mm"):
        print("\nWood selection incomplete. Use Wood Menu to set species and size.")
        return False
    
    if not wood_selection.get("event"):
        print("\nEvent not selected. Use Wood Menu -> Select event (SB/UH) or Main Menu option 3.")
        return False
    
    return True

    # Calculate and display AI-enhanced handicap marks for the heat

def view_handicaps(heat_assignment_df, wood_selection):
    if heat_assignment_df.empty:
        print("No competitors in heat assignment.")
        return

    event_code = wood_selection.get("event")
    if event_code not in ("SB", "UH"):
        print("Invalid or missing event code. Use Wood Menu to select SB or UH.")
        return

    # Load historical results data
    results_df = load_results_df()
    
    if results_df.empty:
        print("\nNo historical data available. Cannot generate AI predictions.")
        print("Please ensure competitors have historical times entered in the Results sheet.")
        return

    # Get wood selection parameters with defaults
    species = wood_selection.get("species", "Unknown")
    diameter = wood_selection.get("size_mm", 300)
    quality = wood_selection.get("quality", 5)
    
    # Ensure quality is an integer
    if quality is None:
        quality = 5
    
    # Calculate AI-enhanced handicaps
    results = calculate_ai_enhanced_handicaps(
        heat_assignment_df,
        species,
        diameter,
        quality,
        event_code,
        results_df
    )
    
    if not results:
        print("\nUnable to generate handicap marks. Please check historical data.")
        return

    # Display compact results
    print("\n" + "="*70)
    print("CALCULATED HANDICAP MARKS")
    print("="*70)
    
    for result in results:
        #pull explanation gnerated from model
        explanation_text = result['explanation']
        # Find the part after the predicted time
        if '(' in explanation_text:
            data_info = explanation_text[explanation_text.find('(')+1:]
            if ')' in data_info:
                data_info = data_info[:data_info.rfind(')')]
            else:
                data_info = explanation_text
        else:
            data_info = explanation_text
        
        print(f"{result['name']:25s} Mark {result['mark']:3d}  ({result['predicted_time']:.1f}s predicted) ({data_info}) [Confidence: {result['confidence']}]")
    
    # Display wood selection
    print("\n" + "="*70)
    print(f"Selected Wood -> Species: {species}, " 
          f"Diameter: {diameter} mm, "
          f"Quality: {quality}")
    print(f"Event: {event_code}")
    print("="*70)
    
    # Offer Monte Carlo simulation
    print("\nWould you like to run a Monte Carlo simulation to validate these handicaps?")
    print("This will simulate 250,000 races to assess fairness.")
    choice = input("Run simulation? (y/n): ").strip().lower()
    
    if choice == 'y':
        simulate_and_assess_handicaps(results, num_simulations=250000)



## MENU OPTION 5: UPDATE RESULTS WITH HEAT RESULTS

'''This menus wiull allow a judge to append the "Results" tab of the 'woodchopping' excel sheet.
    The idea is that after a heat, the judge can update the results and the next heat a competitor races, the handicap will be updated and a more accurate 
     mark will be assigned (can be used to help order Finals) '''

#Find the excel file and sheet names htat will be used to store results
def detect_results_sheet(wb):
    if RESULTS_SHEET in wb.sheetnames:
        return wb[RESULTS_SHEET]
    
    ws = wb.create_sheet(RESULTS_SHEET)
    ws.append(["CompetitorID", "Event", "Time (seconds)", "Size (mm)", "Species Code", "Date"])
    return ws

#Add the results to the 'Results' sheet in the Excel file
'''Prompts the user for: Event (SB/UH), species, wood, quality, and notes (could be heat ID or notes about finals/qualifier or simply the date)'''

def append_results_to_excel(heat_assignment_df, wood_selection, round_object=None, tournament_state=None):
    """Append heat results to Excel Results sheet.

    Args:
        heat_assignment_df (DataFrame): LEGACY - competitors in heat (for backward compatibility)
        wood_selection (dict): Wood characteristics
        round_object (dict): NEW - Round object from tournament system (optional)
        tournament_state (dict): NEW - Tournament state for context (optional)
    """

    # Determine if using new tournament system or legacy single-heat system
    if round_object is not None:
        # NEW TOURNAMENT SYSTEM
        competitors_list = round_object['competitors']
        round_name = round_object['round_name']
    else:
        # LEGACY SINGLE-HEAT SYSTEM
        if heat_assignment_df is None or heat_assignment_df.empty:
            print("No competitors in heat assignment.")
            return
        competitors_list = heat_assignment_df['competitor_name'].tolist()
        round_name = None

    event_code = wood_selection.get("event")
    if event_code not in ("SB", "UH"):
        print("Event not selected. Use Wood Menu -> Select event (SB/UH) or Main Menu option 3.")
        return

    species = wood_selection.get("species")
    size_mm = wood_selection.get("size_mm")
    quality = wood_selection.get("quality")

    # Generate HeatID
    if round_object and tournament_state:
        # NEW: Use round name and tournament context
        event_name = tournament_state.get('event_name', 'Event')
        heat_id = f"{event_code}-{event_name}-{round_name}".replace(" ", "-")
    else:
        # LEGACY: Prompt for Heat ID
        heat_id = input("Enter a Heat ID (e.g., SB-01-Qual or any short label): ").strip()
        if not heat_id:
            heat_id = f"{event_code}-{datetime.now().strftime('%Y%m%d-%H%M%S')}"

    # Get ID/name mapping
    _, name_to_id = get_competitor_id_name_mapping()

    rows_to_write = []
    times_collected = {}

    # STEP 1: Collect raw cutting times (for historical data)
    print("\n" + "=" * 70)
    print("STEP 1: RECORD CUTTING TIMES")
    print("=" * 70)
    print("Enter the raw cutting time for each competitor (from their mark to block severed)")
    print("Press Enter to skip a competitor\n")

    for name in competitors_list:
        s = input(f"  Cutting time for {name}: ").strip()

        if s == "":
            continue

        try:
            t = float(s)
            times_collected[name] = t
        except ValueError:
            print("    Invalid time; skipping this entry.")
            continue

    if not times_collected:
        print("\nNo results to record.")
        return

    # STEP 2: Record finish order (WHO FINISHED FIRST in real-time)
    print("\n" + "=" * 70)
    print("STEP 2: RECORD FINISH ORDER")
    print("=" * 70)
    print("Enter the finish position for each competitor (1 = finished first, 2 = second, etc.)")
    print("This is based on when they physically severed their block (handicap delays included)\n")

    finish_order = {}
    for name in times_collected.keys():
        while True:
            pos_str = input(f"  Finish position for {name}: ").strip()
            try:
                position = int(pos_str)
                if position < 1:
                    print("    Position must be 1 or greater")
                    continue
                finish_order[name] = position
                break
            except ValueError:
                print("    Invalid position; please enter a number")

    # Prepare rows for Excel and update round_object
    timestamp = datetime.now().isoformat(timespec="seconds")

    for name, time_val in times_collected.items():
        competitor_id = name_to_id.get(str(name).strip().lower(), name)

        # Include Round and HeatID columns
        rows_to_write.append([competitor_id, event_code, time_val, size_mm, species, timestamp, round_name or "", heat_id])

        # Store in round_object if using tournament system
        if round_object is not None:
            round_object['actual_results'][name] = time_val
            if 'finish_order' not in round_object:
                round_object['finish_order'] = {}
            round_object['finish_order'][name] = finish_order.get(name, 999)

    if not rows_to_write:
        print("No results to write.")
        return

    try:
        # Load or create workbook
        try:
            wb = load_workbook(RESULTS_FILE)
        except Exception:
            wb = Workbook()
            if "Sheet" in wb.sheetnames and RESULTS_SHEET not in wb.sheetnames:
                wb.remove(wb["Sheet"])

        ws = detect_results_sheet(wb)

        # Ensure header exists with Excel column names (NEW: includes Round and HeatID)
        if ws.max_row == 0:
            ws.append(["CompetitorID", "Event", "Time (seconds)", "Size (mm)", "Species Code", "Date", "Round", "HeatID"])

        # Append rows
        for r in rows_to_write:
            ws.append(r)

        wb.save(RESULTS_FILE)
        wb.close()
        print("Results appended to Excel successfully.")

        # NEW: Update round status if using tournament system
        if round_object is not None:
            round_object['status'] = 'in_progress'  # Mark as in progress (not completed until advancers selected)

    except Exception as e:
        print(f"Error appending results: {e}")