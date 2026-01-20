"""Handicap display and results recording UI functions.

This module handles handicap viewing and results operations including:
- Viewing handicap marks with Monte Carlo validation
- Manual handicap adjustment with judge approval
- Validating heat data completeness
- Recording and saving heat results to Excel
"""

from datetime import datetime
from typing import Dict, Optional, List, Tuple
import pandas as pd
from openpyxl import load_workbook, Workbook

from woodchopping.data import (
    load_results_df,
    get_competitor_id_name_mapping,
    detect_results_sheet
)
from woodchopping.data.validation import (
    is_high_variance_diameter,
    get_diameter_variance_warning,
    check_diameter_sample_size
)
from woodchopping.handicaps import calculate_ai_enhanced_handicaps
from woodchopping.handicaps.qaa_legacy import calculate_qaa_legacy_marks
from woodchopping.simulation import simulate_and_assess_handicaps
from config import paths, rules, sim_config


# File/sheet names from config
RESULTS_FILE = paths.EXCEL_FILE
RESULTS_SHEET = paths.RESULTS_SHEET


def view_handicaps_menu(heat_assignment_df: pd.DataFrame, wood_selection: Dict) -> None:
    """View Handicap Marks Menu.

    Official will be presented with the calculated handicap marks for each
    selected competitor in the heat.

    Menu options:
    1. View handicap marks for current heat
    2. Back to Main Menu

    Args:
        heat_assignment_df: Competitors currently assigned to heat
        wood_selection: Dictionary containing wood characteristics
    """

    # Safety check to make sure that either Standing Block or Underhand is selected
    # If not, will default to None so program doesn't crash
    if "event" not in wood_selection:
        wood_selection["event"] = None

    # Menu loop
    while True:
        print("\n--- View Handicap Marks ---")
        print("1) View handicap marks for current heat")
        print("2) View QAA legacy marks (QAA tables only)")
        print("3) Back to Main Menu")
        s = input("Choose an option: ").strip()

        if s == "1":
            if not validate_heat_data(heat_assignment_df, wood_selection):
                continue
            view_handicaps(heat_assignment_df, wood_selection)
            input("\n(Press Enter to return to the View Handicap Marks menu) ")

        elif s == "2":
            if not validate_heat_data(heat_assignment_df, wood_selection):
                continue
            view_qaa_legacy_handicaps(heat_assignment_df, wood_selection)
            input("\n(Press Enter to return to the View Handicap Marks menu) ")

        elif s == "3" or s == "":
            break

        else:
            print("Invalid selection. Try again.")


def validate_heat_data(heat_assignment_df: pd.DataFrame, wood_selection: Dict) -> bool:
    """Validate that heat data is complete before calculating handicaps.

    Checks specifically for:
    - Competitors in heat assignment
    - Wood species selected
    - Wood size entered
    - Event code selected (SB/UH)

    Args:
        heat_assignment_df: Competitors assigned to heat
        wood_selection: Wood characteristics dictionary

    Returns:
        bool: True if all data is present, False otherwise
    """
    if heat_assignment_df is None or heat_assignment_df.empty:
        print("\nNo competitors in heat assignment. Use Competitor Menu -> Select Competitors for Heat.")
        return False

    if not wood_selection.get("species") or not wood_selection.get("size_mm"):
        print("\nWood selection incomplete. Use Wood Menu to set species and size.")
        return False

    if not wood_selection.get("event"):
        print("\nEvent not selected. Use Wood Menu -> Select event (SB/UH) or Main Menu option 3.")
        return False

    return True


def view_handicaps(heat_assignment_df: pd.DataFrame, wood_selection: Dict) -> None:
    """Calculate and display AI-enhanced handicap marks for the heat.

    This function:
    1. Validates event code
    2. Loads historical results data
    3. Calculates AI-enhanced handicaps
    4. Displays handicap marks with predictions and confidence levels
    5. Optionally runs Monte Carlo simulation for fairness validation

    Args:
        heat_assignment_df: Competitors in heat
        wood_selection: Wood characteristics and event code
    """
    if heat_assignment_df.empty:
        print("No competitors in heat assignment.")
        return

    event_code = wood_selection.get("event")
    if event_code not in ("SB", "UH"):
        print("Invalid or missing event code. Use Wood Menu to select SB or UH.")
        return

    # Load historical results data
    results_df = load_results_df()

    if results_df.empty:
        print("\nNo historical data available. Cannot generate AI predictions.")
        print("Please ensure competitors have historical times entered in the Results sheet.")
        return

    # Get wood selection parameters with defaults
    species = wood_selection.get("species", "Unknown")
    diameter = wood_selection.get("size_mm", 300)
    quality = wood_selection.get("quality", 5)

    # Ensure quality is an integer
    if quality is None:
        quality = 5

    # Calculate AI-enhanced handicaps
    results = calculate_ai_enhanced_handicaps(
        heat_assignment_df,
        species,
        diameter,
        quality,
        event_code,
        results_df
    )

    if not results:
        print("\nUnable to generate handicap marks. Please check historical data.")
        return

    def _truncate_text(text: str, max_len: int) -> str:
        if max_len <= 0:
            return ""
        if len(text) <= max_len:
            return text
        if max_len <= 3:
            return text[:max_len]
        return text[: max_len - 3] + "..."

    # Display compact results
    print("\n" + "=" * 70)
    print("CALCULATED HANDICAP MARKS")
    print("=" * 70)
    print(f"{'Competitor':<28} {'Mark':>4} {'Pred':>7} {'Conf':<8} {'Method':<10}")
    print("-" * 70)

    for result in results:
        name = _truncate_text(result['name'], 28)
        mark = result['mark']
        predicted_time = result['predicted_time']
        confidence = _truncate_text(str(result.get('confidence', 'N/A')), 8)
        method_used = _truncate_text(str(result.get('method_used', 'N/A')), 10)

        print(f"{name:<28} {mark:>4d} {predicted_time:>6.1f}s {confidence:<8} {method_used:<10}")

        explanation_text = result.get('explanation', '')
        data_info = ""
        if '(' in explanation_text:
            data_info = explanation_text[explanation_text.find('(') + 1:]
            if ')' in data_info:
                data_info = data_info[:data_info.rfind(')')]
        if data_info:
            data_info = _truncate_text(data_info, 64)
            print(f"  Data: {data_info}")

    # Display wood selection
    print("\n" + "="*70)
    print(f"Selected Wood -> Species: {species}, "
          f"Diameter: {diameter} mm, "
          f"Quality: {quality}")
    print(f"Event: {event_code}")
    print("="*70)

    # Display warnings for high-variance diameters
    if is_high_variance_diameter(diameter):
        warning_msg = get_diameter_variance_warning(diameter)
        if warning_msg:
            print(f"\n{warning_msg}")
            input("\nPress Enter to continue...")

    # Check sample size for this diameter
    sample_count, confidence = check_diameter_sample_size(results_df, diameter, event_code)
    if confidence in ["VERY LOW", "LOW"]:
        print(f"\n{'='*70}")
        print(f"  SPARSE DATA WARNING")
        print(f"{'='*70}")
        print(f"Historical data for {int(diameter)}mm {event_code}: {sample_count} results")
        print(f"Confidence level: {confidence}")
        print(f"\nHandicaps based on LIMITED historical data - expect higher uncertainty.")
        print(f"{'='*70}")
        input("\nPress Enter to continue...")

    # Offer Monte Carlo simulation
    print("\nWould you like to run a Monte Carlo simulation to validate these handicaps?")
    print(f"This will simulate {sim_config.NUM_SIMULATIONS:,} races to assess fairness.")
    choice = input("Run simulation? (y/n): ").strip().lower()

    if choice == 'y':
        simulate_and_assess_handicaps(results, num_simulations=sim_config.NUM_SIMULATIONS)


def view_qaa_legacy_handicaps(heat_assignment_df: pd.DataFrame, wood_selection: Dict) -> None:
    """Display QAA legacy marks for the current heat."""
    if heat_assignment_df.empty:
        print("No competitors in heat assignment.")
        return

    event_code = wood_selection.get("event")
    if event_code not in ("SB", "UH"):
        print("QAA legacy marks are only available for SB/UH events.")
        return

    results_df = load_results_df()
    if results_df.empty:
        print("\nNo historical data available. Cannot generate QAA legacy marks.")
        return

    species = wood_selection.get("species", "Unknown")
    diameter = wood_selection.get("size_mm", 300)
    quality = wood_selection.get("quality", 5)

    results = calculate_qaa_legacy_marks(
        heat_assignment_df,
        species,
        diameter,
        quality,
        event_code,
        results_df
    )

    if not results:
        print("\nUnable to generate QAA legacy marks.")
        return

    results.sort(key=lambda x: x['mark'])

    print("\n" + "="*70)
    print("QAA LEGACY HANDICAP MARKS (TABLES ONLY)")
    print("="*70)

    for result in results:
        print(
            f"{result['name']:25s} Mark {result['mark']:3d}  "
            f"(300mm book {result['book_mark_300']:.1f}s)  "
            f"{result['explanation']}"
        )

    print("\n" + "="*70)
    print(f"Selected Wood -> Species: {species}, Diameter: {diameter} mm, Quality: {quality}")
    print(f"Event: {event_code}")
    print("="*70)


def _display_live_standings(times_collected: Dict[str, float],
                           all_competitors: List[str],
                           num_to_advance: Optional[int] = None) -> None:
    """
    Display live standings as results are entered (A4 Feature).

    Args:
        times_collected: Dict mapping competitor names to times (so far)
        all_competitors: List of all competitors in the round
        num_to_advance: Number of competitors who advance (None for finals)
    """
    print("\n" + "-" * 70)
    print("Current Standings".center(70))
    print("+" + "-" * 5 + "+" + "-" * 35 + "+" + "-" * 12 + "+" + "-" * 10 + "+")
    print("| Pos | Competitor                        | Time       | Status   |")
    print("+" + "-" * 5 + "+" + "-" * 35 + "+" + "-" * 12 + "+" + "-" * 10 + "+")

    # Sort by time
    sorted_results = sorted(times_collected.items(), key=lambda x: x[1])

    # Determine statuses
    for rank, (name, time_val) in enumerate(sorted_results, 1):
        name_str = name[:33].ljust(33)
        time_str = f"{time_val:.1f} sec".center(10)

        # Determine status
        if num_to_advance is None or num_to_advance == 0:
            # Final round - no advancement
            status = "  --  "
        else:
            # Not all results in yet
            total_entered = len(times_collected)
            total_competitors = len(all_competitors)

            if total_entered < total_competitors:
                # Still waiting for results
                if rank <= num_to_advance:
                    status = "   [OK]   "  # Currently advancing
                elif rank == num_to_advance + 1:
                    status = "   ?   "  # On the bubble
                else:
                    status = "   ?   "  # TBD
            else:
                # All results in
                if rank <= num_to_advance:
                    status = "   [OK]   "  # Advancing
                else:
                    status = "   ?   "  # Eliminated

        # Highlight bubble position
        bubble_marker = " *" if (num_to_advance and rank == num_to_advance) else "  "

        print(f"| {rank:2d}{bubble_marker} | {name_str} | {time_str} | {status} |")

    print("+" + "-" * 5 + "+" + "-" * 35 + "+" + "-" * 12 + "+" + "-" * 10 + "+")

    # Show pending results
    pending = [c for c in all_competitors if c not in times_collected]
    if pending:
        print(f"Pending results: {', '.join(pending)}")

    # Show advancement info
    if num_to_advance and num_to_advance > 0:
        print(f"\nTop {num_to_advance} advance to next round")
        if len(times_collected) == len(all_competitors):
            print("[OK] All results entered - standings are final")

    print("-" * 70 + "\n")


def append_results_to_excel(heat_assignment_df: Optional[pd.DataFrame] = None,
                           wood_selection: Optional[Dict] = None,
                           round_object: Optional[Dict] = None,
                           tournament_state: Optional[Dict] = None) -> None:
    """Append heat results to Excel Results sheet.

    This function supports both:
    - NEW TOURNAMENT SYSTEM: Uses round_object and tournament_state
    - LEGACY SINGLE-HEAT SYSTEM: Uses heat_assignment_df

    Collects:
    1. Raw cutting times (for historical data)
    2. Finish order (who finished first in real-time, including handicap)

    Args:
        heat_assignment_df: LEGACY - competitors in heat (for backward compatibility)
        wood_selection: Wood characteristics
        round_object: NEW - Round object from tournament system (optional)
        tournament_state: NEW - Tournament state for context (optional)
    """

    # Determine if using new tournament system or legacy single-heat system
    if round_object is not None:
        # NEW TOURNAMENT SYSTEM
        competitors_list = round_object['competitors']
        round_name = round_object['round_name']
    else:
        # LEGACY SINGLE-HEAT SYSTEM
        if heat_assignment_df is None or heat_assignment_df.empty:
            print("No competitors in heat assignment.")
            return
        competitors_list = heat_assignment_df['competitor_name'].tolist()
        round_name = None

    event_code = wood_selection.get("event")
    if event_code not in ("SB", "UH"):
        print("Event not selected. Use Wood Menu -> Select event (SB/UH) or Main Menu option 3.")
        return

    species = wood_selection.get("species")
    size_mm = wood_selection.get("size_mm")
    quality = wood_selection.get("quality")

    # Generate HeatID
    if round_object and tournament_state:
        # NEW: Use round name and tournament context
        event_name = tournament_state.get('event_name', 'Event')
        heat_id = f"{event_code}-{event_name}-{round_name}".replace(" ", "-")
    else:
        # LEGACY: Prompt for Heat ID
        heat_id = input("Enter a Heat ID (e.g., SB-01-Qual or any short label): ").strip()
        if not heat_id:
            heat_id = f"{event_code}-{datetime.now().strftime('%Y%m%d-%H%M%S')}"

    # Get ID/name mapping
    _, name_to_id = get_competitor_id_name_mapping()

    rows_to_write = []
    times_collected = {}

    def _build_mark_map() -> Dict[str, int]:
        if round_object and round_object.get('handicap_results'):
            return {r.get('name'): r.get('mark') for r in round_object['handicap_results']}
        if heat_assignment_df is not None and not heat_assignment_df.empty:
            if 'competitor_name' in heat_assignment_df.columns and 'mark' in heat_assignment_df.columns:
                return dict(zip(heat_assignment_df['competitor_name'], heat_assignment_df['mark']))
        return {}

    def _collect_finish_order() -> Dict[str, int]:
        print("\n" + "=" * 70)
        print("RECORD FINISH ORDER")
        print("=" * 70)
        print("Enter the finish position for each competitor (1 = finished first, 2 = second, etc.)")
        print("This is based on when they physically severed their block (handicap delays included)\n")
        finish_order_local = {}
        for name in competitors_list:
            while True:
                pos_str = input(f"  Finish position for {name}: ").strip()
                try:
                    position = int(pos_str)
                    if position < 1:
                        print("    Position must be 1 or greater")
                        continue
                    finish_order_local[name] = position
                    break
                except ValueError:
                    print("    Invalid position; please enter a number")
        return finish_order_local

    def _collect_times(require_all: bool) -> Dict[str, float]:
        print("\n" + "=" * 70)
        print("RECORD CUTTING TIMES (Live Standings)")
        print("=" * 70)
        print("Enter the raw cutting time for each competitor (from their mark to block severed)")
        print("Standings will update after each entry")
        if not require_all:
            print("Press Enter to skip a competitor")
        print("Type 'edit' after entry to adjust a time by name")
        print("")

        # Determine advancement threshold if this is not a final
        num_to_advance = None
        if round_object:
            num_to_advance = round_object.get('num_to_advance', 0)

        times_local = {}
        for idx, name in enumerate(competitors_list, 1):
            while True:
                s = input(f"  [{idx}/{len(competitors_list)}] Cutting time for {name}: ").strip()
                if s.lower() == "edit":
                    if not times_local:
                        print("    No times entered yet.")
                        continue
                    _edit_times(times_local, competitors_list, num_to_advance)
                    continue
                if s == "":
                    if require_all:
                        print("    Time is required for time-only mode.")
                        continue
                    break
                try:
                    t = float(s)
                    times_local[name] = t
                    if len(times_local) > 0:
                        _display_live_standings(times_local, competitors_list, num_to_advance)
                    break
                except ValueError:
                    print("    Invalid time; please enter a number.")
            if s == "" and not require_all:
                continue
        return times_local

    def _edit_times(times_local: Dict[str, float],
                    all_competitors: List[str],
                    num_to_advance: Optional[int]) -> None:
        print("\nEdit a time (press Enter to stop).")
        while True:
            name = input("  Competitor name to edit: ").strip()
            if name == "":
                break
            matched = [n for n in times_local.keys() if n.lower() == name.lower()]
            if not matched:
                print("    Name not found in entered times.")
                continue
            target = matched[0]
            while True:
                s = input(f"  New time for {target}: ").strip()
                try:
                    times_local[target] = float(s)
                    print(f"    Updated {target} to {times_local[target]:.2f}s")
                    if len(times_local) > 0:
                        _display_live_standings(times_local, all_competitors, num_to_advance)
                    break
                except ValueError:
                    print("    Invalid time; please enter a number.")

    def _compute_finish_order_from_times(times_local: Dict[str, float]) -> Dict[str, int]:
        mark_map = _build_mark_map()
        if not mark_map:
            print("Note: No handicap marks found; placings based on raw times.")
        finish_times = []
        for name, time_val in times_local.items():
            mark = mark_map.get(name)
            finish_time = time_val if mark is None else (time_val + float(mark))
            finish_times.append((name, finish_time))
        finish_times.sort(key=lambda x: x[1])
        return {name: idx + 1 for idx, (name, _) in enumerate(finish_times)}

    print("\n" + "=" * 70)
    print("RECORD RESULTS")
    print("=" * 70)
    print("Choose how to enter results:")
    print("1) Placings only (fastest)")
    print("2) Times only (auto-calc placings from handicap + time)")
    print("3) Placings + times")
    choice = input("Selection (1/2/3): ").strip()

    if choice == "1":
        finish_order = _collect_finish_order()
        if round_object is not None:
            if 'finish_order' not in round_object:
                round_object['finish_order'] = {}
            round_object['finish_order'].update(finish_order)
        print("\nPlacings recorded. No times saved to Excel.")
        return

    if choice == "2":
        times_collected = _collect_times(require_all=True)
        finish_order = _compute_finish_order_from_times(times_collected)
        if round_object is not None:
            if 'finish_order' not in round_object:
                round_object['finish_order'] = {}
            round_object['finish_order'].update(finish_order)
        if round_object is not None:
            if 'actual_results' not in round_object:
                round_object['actual_results'] = {}
            round_object['actual_results'].update(times_collected)
    else:
        finish_order = _collect_finish_order()
        if round_object is not None:
            if 'finish_order' not in round_object:
                round_object['finish_order'] = {}
            round_object['finish_order'].update(finish_order)
        times_collected = _collect_times(require_all=False)
        if round_object is not None and times_collected:
            if 'actual_results' not in round_object:
                round_object['actual_results'] = {}
            round_object['actual_results'].update(times_collected)

    if not times_collected:
        print("\nNo cutting times recorded. Finish order saved; nothing written to Excel.")
        return

    # Prepare rows for Excel and update round_object
    timestamp = datetime.now().isoformat(timespec="seconds")

    for name, time_val in times_collected.items():
        competitor_id = name_to_id.get(str(name).strip().lower(), name)

        # Include Round and HeatID columns
        rows_to_write.append([competitor_id, event_code, time_val, size_mm, species, timestamp, round_name or "", heat_id])

        # Store in round_object if using tournament system
        if round_object is not None:
            if 'actual_results' not in round_object:
                round_object['actual_results'] = {}
            round_object['actual_results'][name] = time_val
            if 'finish_order' not in round_object:
                round_object['finish_order'] = {}
            round_object['finish_order'][name] = finish_order.get(name, 999)

    if not rows_to_write:
        print("No results to write.")
        return

    try:
        # Load or create workbook
        try:
            wb = load_workbook(RESULTS_FILE)
        except Exception:
            wb = Workbook()
            if "Sheet" in wb.sheetnames and RESULTS_SHEET not in wb.sheetnames:
                wb.remove(wb["Sheet"])

        ws = detect_results_sheet(wb)

        # Ensure header exists with Excel column names (NEW: includes Round and HeatID)
        if ws.max_row == 0:
            ws.append(["CompetitorID", "Event", "Time (seconds)", "Size (mm)", "Species Code", "Date", "Round", "HeatID"])

        # Append rows
        for r in rows_to_write:
            ws.append(r)

        wb.save(RESULTS_FILE)
        wb.close()
        print("Results appended to Excel successfully.")

        # NEW: Update round status if using tournament system
        if round_object is not None:
            round_object['status'] = 'in_progress'  # Mark as in progress (not completed until advancers selected)

    except Exception as e:
        print(f"Error appending results: {e}")


def judge_approval() -> Tuple[str, str]:
    """
    Prompt judge for approval initials and capture timestamp.

    Returns:
        Tuple of (judge_initials, timestamp_str)
        Returns ('', '') if judge declines approval

    Example:
        >>> initials, timestamp = judge_approval()
        >>> if initials:
        ...     print(f"Approved by {initials} at {timestamp}")
    """
    print("\n" + "="*70)
    print("  JUDGE APPROVAL REQUIRED")
    print("="*70)
    print("\nBy entering your initials, you certify that:")
    print("  - These handicap marks are fair and appropriate")
    print("  - All calculations have been reviewed")
    print("  - Marks comply with AAA competition rules")
    print()

    while True:
        initials = input("Enter your initials to approve (or 'cancel' to abort): ").strip().upper()

        if initials.lower() == 'cancel':
            print("\nApproval cancelled. Handicaps NOT saved.")
            return '', ''

        if len(initials) < 2 or len(initials) > 5:
            print("ERROR: Initials must be 2-5 characters")
            continue

        if not initials.replace('.', '').isalpha():
            print("ERROR: Initials must contain only letters (and optional periods)")
            continue

        # Confirm
        confirm = input(f"\nConfirm approval by {initials}? (y/n): ").strip().lower()
        if confirm == 'y':
            timestamp = datetime.now().isoformat(timespec="seconds")
            print(f"\n[OK] Approved by {initials} at {timestamp}")
            return initials, timestamp
        else:
            print("Approval cancelled. Please re-enter initials.")


def manual_adjust_handicaps(
    handicap_results: List[Dict],
    wood_selection: Dict
) -> Tuple[List[Dict], str, str]:
    """
    Allow judge to manually adjust individual handicap marks with validation.

    This function:
    1. Displays current handicap list
    2. Allows selection of competitor to adjust
    3. Validates new mark (3 <= mark <= 180 per AAA rules)
    4. Shows updated handicap sheet after each change
    5. Repeats until judge is satisfied

    Args:
        handicap_results: List of handicap dicts with 'name', 'mark', 'predicted_time', etc.
        wood_selection: Wood characteristics for display context

    Returns:
        Tuple of (adjusted_handicaps, judge_initials, timestamp)
        Returns (original_handicaps, '', '') if no changes made or approval cancelled

    AAA Rules Enforced:
        - Minimum mark: 3 seconds
        - Maximum mark: 180 seconds
        - Marks must be whole seconds (integers)

    Example:
        >>> adjusted, initials, timestamp = manual_adjust_handicaps(results, wood_sel)
        >>> if initials:
        ...     print(f"Handicaps adjusted and approved by {initials}")
    """
    # Work with a copy to avoid modifying original
    adjusted_results = [result.copy() for result in handicap_results]

    print("\n" + "="*70)
    print("  MANUAL HANDICAP ADJUSTMENT")
    print("="*70)
    print("\nYou can manually override calculated handicap marks.")
    print("This is useful when:")
    print("  - Competitor has recent injury or condition affecting performance")
    print("  - Historical data doesn't reflect current ability")
    print("  - Judge has specific knowledge about competitor")
    print()

    while True:
        # Display current handicaps
        _display_handicap_summary(adjusted_results, wood_selection)

        print("\n" + "="*70)
        print("OPTIONS:")
        print("  1-{}  Select competitor number to adjust".format(len(adjusted_results)))
        print("  'done' Accept current handicaps and proceed to approval")
        print("  'cancel' Abandon changes and return to original handicaps")
        print("="*70)

        choice = input("\nYour choice: ").strip().lower()

        if choice == 'done':
            # Proceed to approval
            print("\n" + "="*70)
            print("FINAL HANDICAP MARKS:")
            _display_handicap_summary(adjusted_results, wood_selection)

            initials, timestamp = judge_approval()
            if initials:
                return adjusted_results, initials, timestamp
            else:
                # Approval cancelled - ask if they want to continue adjusting
                retry = input("\nContinue adjusting? (y/n): ").strip().lower()
                if retry != 'y':
                    return handicap_results, '', ''  # Return original unchanged
                continue

        elif choice == 'cancel':
            confirm = input("\nAre you sure you want to abandon all changes? (y/n): ").strip().lower()
            if confirm == 'y':
                print("Changes abandoned. Returning to original handicaps.")
                return handicap_results, '', ''
            continue

        else:
            # Try to parse as competitor number
            try:
                comp_num = int(choice)
                if comp_num < 1 or comp_num > len(adjusted_results):
                    print(f"ERROR: Invalid number. Choose 1-{len(adjusted_results)}")
                    continue

                # Get the competitor
                comp_idx = comp_num - 1
                competitor = adjusted_results[comp_idx]

                # Show current mark
                print(f"\n{competitor['name']}")
                print(f"  Current mark: {competitor['mark']} seconds")
                print(f"  Predicted time: {competitor['predicted_time']:.1f}s")
                print(f"  Confidence: {competitor['confidence']}")

                # Get new mark
                while True:
                    new_mark_str = input(f"\nEnter new mark ({rules.MIN_MARK_SECONDS}-{rules.MAX_TIME_LIMIT_SECONDS}s, or 'cancel'): ").strip()

                    if new_mark_str.lower() == 'cancel':
                        print("Adjustment cancelled for this competitor.")
                        break

                    try:
                        new_mark = int(new_mark_str)

                        # Validate against AAA rules
                        if new_mark < rules.MIN_MARK_SECONDS:
                            print(f"ERROR: Mark must be at least {rules.MIN_MARK_SECONDS} seconds (AAA rule)")
                            continue

                        if new_mark > rules.MAX_TIME_LIMIT_SECONDS:
                            print(f"ERROR: Mark cannot exceed {rules.MAX_TIME_LIMIT_SECONDS} seconds (AAA rule)")
                            continue

                        # Show change
                        old_mark = competitor['mark']

                        # Prompt for reason (A5 feature)
                        print("\n" + "-" * 70)
                        print("Please explain why you're adjusting this handicap (for audit trail):")
                        reason = input("Reason: ").strip()
                        while not reason:
                            print("[WARN] Reason is required for adjustment tracking.")
                            reason = input("Reason: ").strip()

                        competitor['mark'] = new_mark
                        competitor['manual_adjustment'] = True
                        competitor['original_mark'] = old_mark
                        competitor['adjustment_reason'] = reason  # A5: Store reason

                        print(f"\n[OK] Updated {competitor['name']}: Mark {old_mark} -> {new_mark}")
                        print(f"  Reason: {reason}")
                        break

                    except ValueError:
                        print("ERROR: Mark must be a whole number")
                        continue

            except ValueError:
                print(f"ERROR: Invalid input. Enter number (1-{len(adjusted_results)}), 'done', or 'cancel'")
                continue


def _display_handicap_summary(handicap_results: List[Dict], wood_selection: Dict) -> None:
    """
    Display summary table of handicap marks (helper function).

    Args:
        handicap_results: List of handicap dicts
        wood_selection: Wood characteristics for context
    """
    print("\n" + "="*70)
    print("HANDICAP MARKS")
    print("="*70)

    print(f"\n{'#':<4} {'Competitor':<25} {'Mark':<6} {'Pred Time':<10} {'Confidence':<12} {'Status'}")
    print("-"*70)

    for idx, result in enumerate(handicap_results, 1):
        status = "*ADJUSTED*" if result.get('manual_adjustment') else ""
        print(f"{idx:<4} {result['name']:<25} {result['mark']:<6} {result['predicted_time']:<10.1f} {result['confidence']:<12} {status}")

    print("\n" + "="*70)

    # Get species name from code
    from woodchopping.data import get_species_name_from_code
    species_code = wood_selection.get('species', 'Unknown')
    species_name = get_species_name_from_code(species_code)

    print(f"Wood: {species_name}, "
          f"{wood_selection.get('size_mm', '?')}mm, "
          f"Quality {wood_selection.get('quality', '?')}")
    print(f"Event: {wood_selection.get('event', '?')}")
    print("="*70)
