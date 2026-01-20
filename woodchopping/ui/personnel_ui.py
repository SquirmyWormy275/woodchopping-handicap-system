"""Personnel management UI functions.

This module handles roster management operations including:
- Adding new competitors with historical times
- Viewing full roster
- Managing historical time entries
"""

from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from woodchopping.data import (
    load_competitors_df,
    get_competitor_id_name_mapping,
    save_time_to_results
)
from config import paths


# File/sheet names from config
COMPETITOR_FILE = paths.EXCEL_FILE
COMPETITOR_SHEET = paths.COMPETITOR_SHEET
RESULTS_FILE = paths.EXCEL_FILE
RESULTS_SHEET = paths.RESULTS_SHEET


def personnel_management_menu(comp_df: pd.DataFrame) -> pd.DataFrame:
    """Personnel Management Menu - for adding/editing/removing competitors from master roster.

    This is separated from tournament operations to allow roster management
    before or during tournament setup.

    Args:
        comp_df: Current competitor roster

    Returns:
        DataFrame: Updated competitor roster
    """

    while True:
        print("\n" + "=" * 60)
        print("  PERSONNEL MANAGEMENT MENU")
        print("=" * 60)
        print("1) Add New Competitor to Roster")
        print("2) View Full Roster")
        print("3) Remove Competitor from Roster")
        print("4) Back to Main Menu")
        print("=" * 60)

        choice = input("Choose an option: ").strip()

        if choice == "1":
            # Add new competitor with historical times
            comp_df = add_competitor_with_times()
            print("\n[OK] Competitor added successfully")

        elif choice == "2":
            # View full roster
            if comp_df is None or comp_df.empty:
                print("\nRoster is empty. No competitors found.")
            else:
                print(f"\n{'='*60}")
                print(f"  FULL ROSTER ({len(comp_df)} competitors)")
                print(f"{'='*60}")
                for idx in range(len(comp_df)):
                    row = comp_df.iloc[idx]
                    name = row.get("competitor_name", "Unknown")
                    country = row.get("competitor_country", "Unknown")
                    comp_id = row.get("competitor_id", "N/A")
                    print(f"{idx + 1:3d}) {name:30s} ({country:15s}) [ID: {comp_id}]")
                print(f"{'='*60}")

        elif choice == "3":
            # Remove competitor from roster
            print("\nRemove Competitor - Feature Coming Soon")
            print("(Currently, please remove directly from Excel file)")

        elif choice == "4" or choice == "":
            break

        else:
            print("Invalid selection. Try again.")

    return comp_df


def add_competitor_with_times() -> pd.DataFrame:
    """Add a new competitor to roster and prompt for historical times.

    This function:
    1. Prompts for competitor basic info (name, country, state, gender)
    2. Adds competitor to Excel Competitor sheet
    3. Prompts for historical competition times (minimum 3 required)
    4. Saves times to Results sheet

    Returns:
        DataFrame: Updated competitor roster
    """
    try:
        # Get competitor basic info
        print("\n--- Add New Competitor ---")
        name = input("Enter competitor name: ").strip()
        if not name:
            print("Name cannot be empty.")
            return load_competitors_df()

        country = input("Enter competitor country: ").strip()
        state = input("Enter state/province (optional): ").strip()
        gender = input("Enter gender (M/F, optional): ").strip().upper()

        # Add to competitors sheet
        try:
            wb = load_workbook(COMPETITOR_FILE)
        except FileNotFoundError:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        if COMPETITOR_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(COMPETITOR_SHEET)
            ws.append(["CompetitorID", "Name", "Country", "State/Province", "Gender"])
        else:
            ws = wb[COMPETITOR_SHEET]
            if ws.max_row < 1:
                ws.append(["CompetitorID", "Name", "Country", "State/Province", "Gender"])

        # Check for duplicate
        existing_names = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) > 1 and row[1]:  # Name is in column 2 (index 1)
                existing_names.add(str(row[1]).strip().lower())

        if name.lower() in existing_names:
            print("Competitor already exists in roster.")
            wb.close()
            return load_competitors_df()

        # Generate new CompetitorID
        new_id = f"C{str(ws.max_row).zfill(3)}"

        # Add competitor to sheet
        ws.append([new_id, name, country, state, gender])

        wb.save(COMPETITOR_FILE)
        wb.close()
        print(f"\n[OK] {name} added to roster successfully with ID {new_id}")

        # Now prompt for historical times
        print("\n--- Enter Historical Competition Times ---")
        print("Minimum 3 times required for handicap calculation.")
        print("You can enter more than 3 if desired.")

        add_historical_times_for_competitor(name)

        return load_competitors_df()

    except Exception as e:
        print(f"Error adding competitor: {e}")
        return load_competitors_df()


def add_historical_times_for_competitor(competitor_name: str) -> None:
    """Prompt for and save historical times to results sheet.

    Judge will be prompted to:
    - Select event type (SB or UH)
    - Enter the time in seconds
    - Enter wood species
    - Enter wood diameter in mm
    - Enter wood quality (1-10)
    - Optionally enter date (defaults to current date)

    The program will store this data in the results sheet.

    Args:
        competitor_name: Name of competitor to add times for
    """
    times_added = 0

    while True:
        if times_added < 3:
            print(f"\n--- Historical Time Entry {times_added + 1} (minimum 3 required) ---")
        else:
            cont = input(f"\n{times_added} times entered. Add another? (y/n): ").strip().lower()
            if cont != 'y':
                break
            print(f"\n--- Historical Time Entry {times_added + 1} ---")

        # Get event type
        while True:
            event = input("Event type (SB for Standing Block, UH for Underhand): ").strip().upper()
            if event in ["SB", "UH"]:
                break
            print("Please enter SB or UH.")

        # Get time
        while True:
            time_str = input("Time in seconds (e.g., 45.3): ").strip()
            try:
                time_val = float(time_str)
                break
            except ValueError:
                print("Please enter a valid number.")

        # Get wood species
        species = input("Wood species: ").strip()

        # Get size
        while True:
            size_str = input("Wood diameter in mm: ").strip()
            try:
                size_val = float(size_str)
                break
            except ValueError:
                print("Please enter a valid number.")

        # Get quality (optional, default to 5)
        quality_str = input("Wood quality (1-10, press Enter for 5): ").strip()
        if quality_str:
            try:
                quality = int(quality_str)
                quality = max(1, min(10, quality))
            except:
                quality = 5
        else:
            quality = 5

        # Optional date
        date_str = input("Date (optional, format YYYY-MM-DD or press Enter to skip): ").strip()
        if not date_str:
            timestamp = datetime.now().isoformat(timespec="seconds")
        else:
            try:
                date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                timestamp = date_obj.isoformat(timespec="seconds")
            except:
                timestamp = datetime.now().isoformat(timespec="seconds")

        # Save to results sheet
        save_time_to_results(event, competitor_name, species, size_val, quality, time_val,
                           f"Historical-{event}", timestamp)
        times_added += 1
        print(f"[OK] Time #{times_added} saved successfully")

    if times_added >= 3:
        print(f"\n[OK] {times_added} historical times added for {competitor_name}.")
    else:
        print(f"\n[WARN] Warning: Only {times_added} times added. Minimum 3 recommended for handicap calculation.")
