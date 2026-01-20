"""Excel I/O functions for loading and saving woodchopping competition data."""

import pandas as pd
from openpyxl import load_workbook, Workbook
from typing import Tuple, Dict
from datetime import datetime

# Import config
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
from config import paths


def _resolve_sheet_name(excel_path: str, desired_name: str) -> str:
    """Return the exact sheet name matching desired_name (case-insensitive)."""
    try:
        xl = pd.ExcelFile(excel_path)
        for sheet in xl.sheet_names:
            if sheet.strip().lower() == desired_name.strip().lower():
                return sheet
    except Exception:
        pass
    return desired_name


def get_competitor_id_name_mapping() -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    Load competitor data and return bidirectional ID/name mapping dictionaries.

    Returns:
        Tuple of (id_to_name dict, name_to_id dict)
    """
    try:
        sheet_name = _resolve_sheet_name(paths.EXCEL_FILE, paths.COMPETITOR_SHEET)
        df = pd.read_excel(paths.EXCEL_FILE, sheet_name=sheet_name)

        if df.empty:
            return {}, {}

        id_to_name = {}
        name_to_id = {}

        for _, row in df.iterrows():
            comp_id = str(row.get('CompetitorID', '')).strip()
            name = str(row.get('Name', '')).strip()

            if comp_id and name:
                id_to_name[comp_id] = name
                name_to_id[name.lower()] = comp_id

        return id_to_name, name_to_id

    except Exception as e:
        print(f"Error loading competitor ID/name mapping: {e}")
        return {}, {}


def load_competitors_df() -> pd.DataFrame:
    """
    Load the competitor roster from Excel into a DataFrame.

    Returns:
        DataFrame with competitor information (name, country, ID, state, gender)
    """
    try:
        sheet_name = _resolve_sheet_name(paths.EXCEL_FILE, paths.COMPETITOR_SHEET)
        df = pd.read_excel(paths.EXCEL_FILE, sheet_name=sheet_name)

        # Standardize column names
        column_mapping = {
            'Name': 'competitor_name',
            'Country': 'competitor_country',
            'CompetitorID': 'competitor_id',
            'State/Province': 'state_province',
            'Gender': 'gender'
        }
        df = df.rename(columns=column_mapping)

        if df.empty:
            print("No competitors found in Excel. Please add competitors first.")
            return pd.DataFrame(columns=["competitor_name", "competitor_country"])

        print(f"Roster loaded successfully from Excel. Found {len(df)} competitors.")
        return df

    except FileNotFoundError:
        print(f"Excel file '{paths.EXCEL_FILE}' not found. Creating new file.")
        # Create the Excel file with proper sheets
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        ws = wb.create_sheet(paths.COMPETITOR_SHEET)
        ws.append(["CompetitorID", "Name", "Country", "State/Province", "Gender"])
        wb.save(paths.EXCEL_FILE)
        wb.close()
        return pd.DataFrame(columns=["competitor_name", "competitor_country"])
    except Exception as e:
        print(f"Error loading roster from Excel: {e}")
        print(f"Looking for sheet: {paths.COMPETITOR_SHEET}")
        return pd.DataFrame(columns=["competitor_name", "competitor_country"])


def load_wood_data() -> pd.DataFrame:
    """
    Load wood species data from Excel into a DataFrame.

    Returns:
        DataFrame with wood properties (species, multiplier, janka, etc.)
    """
    try:
        sheet_name = _resolve_sheet_name(paths.EXCEL_FILE, paths.WOOD_SHEET)
        df = pd.read_excel(paths.EXCEL_FILE, sheet_name=sheet_name)
        # Drop unnamed columns from Excel artifacts
        df = df.loc[:, [c for c in df.columns if not str(c).startswith('Unnamed')]].copy()
        df.columns = [str(c).strip() for c in df.columns]

        # Coerce numeric wood property columns
        numeric_cols = ['janka_hard', 'spec_gravity', 'crush_strength', 'shear', 'MOR', 'MOE']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
                if df[col].isna().any():
                    median_val = df[col].median()
                    if pd.notna(median_val):
                        df[col] = df[col].fillna(median_val)

        if 'speciesID' in df.columns:
            df['speciesID'] = df['speciesID'].astype(str).str.strip()

        return df
    except Exception as e:
        print(f"Error loading wood data: {e}")
        return pd.DataFrame(columns=["species", "multiplier"])


# Cache for species name lookups to avoid repeated Excel reads
_species_cache = {}

def get_species_name_from_code(species_code: str) -> str:
    """
    Convert species code (e.g., 'S01') to species name (e.g., 'Eastern White Pine').

    Uses caching to avoid repeated Excel reads.

    Args:
        species_code: Species code like 'S01', 'S02', etc.

    Returns:
        str: Species name, or the code itself if not found
    """
    # Check cache first
    if species_code in _species_cache:
        return _species_cache[species_code]

    try:
        sheet_name = _resolve_sheet_name(paths.EXCEL_FILE, paths.WOOD_SHEET)
        wood_df = pd.read_excel(paths.EXCEL_FILE, sheet_name=sheet_name)

        if 'speciesID' in wood_df.columns and 'species' in wood_df.columns:
            # Build cache for all species
            for _, row in wood_df.iterrows():
                code = row['speciesID']
                name = row['species']
                if pd.notna(code) and pd.notna(name):
                    _species_cache[code] = name

            # Return from cache if found
            if species_code in _species_cache:
                return _species_cache[species_code]

        # Fallback: return the code itself and cache it
        _species_cache[species_code] = species_code
        return species_code

    except Exception as e:
        print(f"Error looking up species name: {e}")
        # Cache the fallback value
        _species_cache[species_code] = species_code
        return species_code


def load_results_df() -> pd.DataFrame:
    """
    Load the Results sheet as a DataFrame with competitor names (not IDs).

    Returns:
        DataFrame with historical results including competitor names
    """
    try:
        # Read raw results
        sheet_name = _resolve_sheet_name(paths.EXCEL_FILE, paths.RESULTS_SHEET)
        df = pd.read_excel(paths.EXCEL_FILE, sheet_name=sheet_name)

        if df.empty:
            print("No results found in Excel.")
            return pd.DataFrame()

        # Get ID to name mapping
        id_to_name, _ = get_competitor_id_name_mapping()

        # Map column names (Phase 5: Added FinishPosition for stacking ensemble)
        column_mapping = {
            'CompetitorID': 'competitor_id',
            'Event': 'event',
            'Time (seconds)': 'raw_time',
            'Size (mm)': 'size_mm',
            'Species Code': 'species',
            'Date': 'date',
            'Date (optional)': 'date',
            'Quality': 'quality',
            'HeatID': 'heat_id',
            'FinishPosition': 'finish_position'  # NEW: Nullable field for stacking ensemble
        }
        df = df.rename(columns=column_mapping)

        # Flexible normalization for variant headers (case-insensitive)
        if 'size_mm' not in df.columns:
            for col in df.columns:
                col_lower = str(col).strip().lower()
                if 'diameter' in col_lower or ('size' in col_lower and 'mm' in col_lower):
                    df = df.rename(columns={col: 'size_mm'})
                    break
        if 'raw_time' not in df.columns:
            for col in df.columns:
                col_lower = str(col).strip().lower()
                if 'time' in col_lower and 'date' not in col_lower:
                    df = df.rename(columns={col: 'raw_time'})
                    break
        if 'event' not in df.columns:
            for col in df.columns:
                col_lower = str(col).strip().lower()
                if 'event' in col_lower:
                    df = df.rename(columns={col: 'event'})
                    break
        if 'species' not in df.columns:
            for col in df.columns:
                col_lower = str(col).strip().lower()
                if 'species' in col_lower:
                    df = df.rename(columns={col: 'species'})
                    break
        if 'competitor_id' not in df.columns and 'competitor_name' not in df.columns:
            for col in df.columns:
                col_lower = str(col).strip().lower()
                if 'competitor' in col_lower and 'id' in col_lower:
                    df = df.rename(columns={col: 'competitor_id'})
                    break
                if col_lower in {'name', 'competitor', 'competitor name'}:
                    df = df.rename(columns={col: 'competitor_name'})
                    break

        # Phase 5: Handle FinishPosition field (backward compatible - nullable)
        if 'finish_position' in df.columns:
            df['finish_position'] = pd.to_numeric(df['finish_position'], errors='coerce')
            # NaN values indicate no finish position recorded (backward compatible)
        else:
            # Add column if missing (backward compatibility)
            df['finish_position'] = pd.NA

        # Convert IDs to names
        if 'competitor_id' in df.columns and 'competitor_name' not in df.columns:
            df['competitor_name'] = df['competitor_id'].apply(
                lambda x: id_to_name.get(str(x).strip(), f"Unknown_{x}")
            )

        # Parse dates to datetime objects (critical for time-decay weighting)
        if 'date' in df.columns:
            df['date'] = pd.to_datetime(df['date'], errors='coerce')
            # Note: errors='coerce' converts invalid/missing dates to NaT (Not a Time)
            # This maintains backward compatibility with results that have no dates

        # Coerce quality to numeric (non-numeric values become NaN)
        if 'quality' in df.columns:
            df['quality'] = pd.to_numeric(df['quality'], errors='coerce')

        # CRITICAL FIX: Normalize event codes to uppercase (fix UH/uh and SB/sb inconsistency)
        # This bug caused separate baselines for uppercase vs lowercase event codes
        if 'event' in df.columns:
            df['event'] = df['event'].str.upper()

        return df

    except FileNotFoundError:
        print(f"Results sheet not found in '{paths.EXCEL_FILE}'.")
        return pd.DataFrame()
    except Exception as e:
        print(f"Error loading results: {e}")
        return pd.DataFrame()


def detect_results_sheet(wb: Workbook):
    """
    Return the Results sheet from workbook; create it with proper headers if missing.

    Args:
        wb: openpyxl Workbook object

    Returns:
        Results worksheet
    """
    if paths.RESULTS_SHEET in wb.sheetnames:
        return wb[paths.RESULTS_SHEET]
    else:
        # Create Results sheet with proper headers
        ws = wb.create_sheet(paths.RESULTS_SHEET)
        ws.append([
            "CompetitorID",
            "Event",
            "Time (seconds)",
            "Size (mm)",
            "Species Code",
            "Quality",
            "HeatID",
            "Date"
        ])
        print(f"Created '{paths.RESULTS_SHEET}' sheet with headers.")
        return ws


def save_time_to_results(
    event: str,
    name: str,
    species: str,
    size: float,
    quality: int,
    time: float,
    heat_id: str,
    timestamp: str
) -> None:
    """
    Save a single time entry to the results sheet.

    Args:
        event: Event code (SB or UH)
        name: Competitor name
        species: Wood species code
        size: Diameter in mm
        quality: Wood quality (1-10)
        time: Time in seconds
        heat_id: Heat identifier
        timestamp: Date/time string
    """
    _, name_to_id = get_competitor_id_name_mapping()

    competitor_id = name_to_id.get(name.lower())
    if not competitor_id:
        print(f"Warning: Could not find ID for {name}")
        return

    # Normalize event code to uppercase
    event = event.upper() if event else event

    try:
        wb = load_workbook(paths.EXCEL_FILE)
        ws = detect_results_sheet(wb)

        ws.append([
            competitor_id,
            event,
            time,
            size,
            species,
            quality,
            heat_id,
            timestamp
        ])

        wb.save(paths.EXCEL_FILE)
        wb.close()

    except Exception as e:
        print(f"Error saving time to results: {e}")


def append_results_to_excel(heat_assignment_df, wood_selection, round_object=None, tournament_state=None, event_name=None):
    """
    Append heat results to Excel Results sheet.

    Supports both legacy single-heat system, single-event tournaments, and multi-event tournaments.

    Args:
        heat_assignment_df (DataFrame): LEGACY - competitors in heat (for backward compatibility)
        wood_selection (dict): Wood characteristics (species, size_mm, quality, event)
        round_object (dict): NEW - Round object from tournament system (optional)
        tournament_state (dict): NEW - Tournament state for single-event context (optional)
        event_name (str): NEW - Event name for multi-event tournaments (optional)
    """
    # Determine if using new tournament system or legacy single-heat system
    if round_object is not None:
        # NEW TOURNAMENT SYSTEM
        competitors_list = round_object['competitors']
        round_name = round_object['round_name']
    else:
        # LEGACY SINGLE-HEAT SYSTEM
        if heat_assignment_df is None or heat_assignment_df.empty:
            print("No competitors in heat assignment.")
            return
        competitors_list = heat_assignment_df['competitor_name'].tolist()
        round_name = None

    event_code = wood_selection.get("event")
    # Normalize event code to uppercase
    if event_code:
        event_code = event_code.upper()
    if event_code not in ("SB", "UH"):
        print("Event not selected. Use Wood Menu -> Select event (SB/UH) or Main Menu option 3.")
        return

    species = wood_selection.get("species")
    size_mm = wood_selection.get("size_mm")
    quality = wood_selection.get("quality")

    # Generate HeatID
    if round_object and (tournament_state or event_name):
        # NEW: Use round name and tournament/event context
        if event_name:
            # Multi-event tournament: use provided event_name
            evt_name = event_name
        elif tournament_state:
            # Single-event tournament: get from tournament_state
            evt_name = tournament_state.get('event_name', 'Event')
        else:
            evt_name = 'Event'

        heat_id = f"{event_code}-{evt_name}-{round_name}".replace(" ", "-")
    else:
        # LEGACY: Prompt for Heat ID
        heat_id = input("Enter a Heat ID (e.g., SB-01-Qual or any short label): ").strip()
        if not heat_id:
            heat_id = f"{event_code}-{datetime.now().strftime('%Y%m%d-%H%M%S')}"

    # Get ID/name mapping
    _, name_to_id = get_competitor_id_name_mapping()

    rows_to_write = []
    times_collected = {}

    def _build_mark_map() -> Dict[str, int]:
        if round_object and round_object.get('handicap_results'):
            return {r.get('name'): r.get('mark') for r in round_object['handicap_results']}
        if heat_assignment_df is not None and not heat_assignment_df.empty:
            if 'competitor_name' in heat_assignment_df.columns and 'mark' in heat_assignment_df.columns:
                return dict(zip(heat_assignment_df['competitor_name'], heat_assignment_df['mark']))
        return {}

    def _collect_finish_order() -> Dict[str, int]:
        print("\n" + "=" * 70)
        print("RECORD FINISH ORDER")
        print("=" * 70)
        print("Enter the finish position for each competitor (1 = finished first, 2 = second, etc.)")
        print("This is based on when they physically severed their block (handicap delays included)\n")
        finish_order_local = {}
        for name in competitors_list:
            while True:
                pos_str = input(f"  Finish position for {name}: ").strip()
                try:
                    position = int(pos_str)
                    if position < 1:
                        print("    Position must be 1 or greater")
                        continue
                    finish_order_local[name] = position
                    break
                except ValueError:
                    print("    Invalid position; please enter a number")
        return finish_order_local

    def _collect_times(require_all: bool) -> Dict[str, float]:
        print("\n" + "=" * 70)
        print("RECORD CUTTING TIMES")
        print("=" * 70)
        print("Enter the raw cutting time for each competitor (from their mark to block severed)")
        if not require_all:
            print("Press Enter to skip a competitor")
        print("Type 'edit' after entry to adjust a time by name")
        print("")
        times_local = {}
        for name in competitors_list:
            while True:
                s = input(f"  Cutting time for {name}: ").strip()
                if s.lower() == "edit":
                    if not times_local:
                        print("    No times entered yet.")
                        continue
                    _edit_times(times_local)
                    continue
                if s == "":
                    if require_all:
                        print("    Time is required for time-only mode.")
                        continue
                    break
                try:
                    times_local[name] = float(s)
                    break
                except ValueError:
                    print("    Invalid time; please enter a number.")
            if s == "" and not require_all:
                continue
        return times_local

    def _edit_times(times_local: Dict[str, float]) -> None:
        print("\nEdit a time (press Enter to stop).")
        while True:
            name = input("  Competitor name to edit: ").strip()
            if name == "":
                break
            matched = [n for n in times_local.keys() if n.lower() == name.lower()]
            if not matched:
                print("    Name not found in entered times.")
                continue
            target = matched[0]
            while True:
                s = input(f"  New time for {target}: ").strip()
                try:
                    times_local[target] = float(s)
                    print(f"    Updated {target} to {times_local[target]:.2f}s")
                    break
                except ValueError:
                    print("    Invalid time; please enter a number.")

    def _compute_finish_order_from_times(times_local: Dict[str, float]) -> Dict[str, int]:
        mark_map = _build_mark_map()
        if not mark_map:
            print("Note: No handicap marks found; placings based on raw times.")
        finish_times = []
        for name, time_val in times_local.items():
            mark = mark_map.get(name)
            finish_time = time_val if mark is None else (time_val + float(mark))
            finish_times.append((name, finish_time))
        finish_times.sort(key=lambda x: x[1])
        return {name: idx + 1 for idx, (name, _) in enumerate(finish_times)}

    print("\n" + "=" * 70)
    print("RECORD RESULTS")
    print("=" * 70)
    print("Choose how to enter results:")
    print("1) Placings only (fastest)")
    print("2) Times only (auto-calc placings from handicap + time)")
    print("3) Placings + times")
    choice = input("Selection (1/2/3): ").strip()

    if choice == "1":
        finish_order = _collect_finish_order()
        if round_object is not None:
            if 'finish_order' not in round_object:
                round_object['finish_order'] = {}
            round_object['finish_order'].update(finish_order)
        print("\nPlacings recorded. No times saved to Excel.")
        return

    if choice == "2":
        times_collected = _collect_times(require_all=True)
        finish_order = _compute_finish_order_from_times(times_collected)
        if round_object is not None:
            if 'finish_order' not in round_object:
                round_object['finish_order'] = {}
            round_object['finish_order'].update(finish_order)
        if round_object is not None:
            if 'actual_results' not in round_object:
                round_object['actual_results'] = {}
            round_object['actual_results'].update(times_collected)
    else:
        finish_order = _collect_finish_order()
        if round_object is not None:
            if 'finish_order' not in round_object:
                round_object['finish_order'] = {}
            round_object['finish_order'].update(finish_order)
        times_collected = _collect_times(require_all=False)
        if round_object is not None and times_collected:
            if 'actual_results' not in round_object:
                round_object['actual_results'] = {}
            round_object['actual_results'].update(times_collected)

    if not times_collected:
        print("\nNo cutting times recorded. Finish order saved; nothing written to Excel.")
        return

    # Prepare rows for Excel and update round_object
    timestamp = datetime.now().isoformat(timespec="seconds")

    for name, time_val in times_collected.items():
        competitor_id = name_to_id.get(str(name).strip().lower(), name)

        # Include Quality and HeatID columns
        rows_to_write.append([
            competitor_id,
            event_code,
            time_val,
            size_mm,
            species,
            quality,
            heat_id,
            timestamp
        ])

        # Store in round_object if using tournament system
        if round_object is not None:
            if 'actual_results' not in round_object:
                round_object['actual_results'] = {}
            round_object['actual_results'][name] = time_val
            if 'finish_order' not in round_object:
                round_object['finish_order'] = {}
            round_object['finish_order'][name] = finish_order.get(name, 999)

    if not rows_to_write:
        print("No results to write.")
        return

    try:
        # Load or create workbook
        try:
            wb = load_workbook(paths.EXCEL_FILE)
        except Exception:
            wb = Workbook()
            if "Sheet" in wb.sheetnames and paths.RESULTS_SHEET not in wb.sheetnames:
                wb.remove(wb["Sheet"])

        ws = detect_results_sheet(wb)

        # Ensure header exists with Excel column names
        if ws.max_row == 0:
            ws.append([
                "CompetitorID",
                "Event",
                "Time (seconds)",
                "Size (mm)",
                "Species Code",
                "Quality",
                "HeatID",
                "Date"
            ])

        # Append rows
        for r in rows_to_write:
            ws.append(r)

        wb.save(paths.EXCEL_FILE)
        wb.close()
        print("Results appended to Excel successfully.")

        # NEW: Update round status if using tournament system
        if round_object is not None:
            round_object['status'] = 'in_progress'  # Mark as in progress (not completed until advancers selected)

    except Exception as e:
        print(f"Error appending results: {e}")
